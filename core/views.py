"""
Core views — الواجهة الأمامية + لوحة المديرة.

التغييرات الرئيسية في هذه النسخة:
- صلاحيات الطالبة: لا تظهر لها بنوك الأسئلة ولا اختبارات التحصيلي ولا
  الاختبار القبلي/البعدي. تظهر لها فقط: مهارات قدرات + دروس تحصيلي
  + الاختبارات الشاملة المفعّلة.
- تفعيل اختبار شامل لـ "جميع الفصول" (classroom_id == 'all').
- مسارات طباعة منفصلة للمحتوى والاختبارات (skill_print, exam_print).
- مسارات استيراد Excel للمعلمات وللطالبات من لوحة المديرة.
- تعديل/حذف الاختبارات الشاملة.
- تعديل بيانات مستخدم (تحديث الاسم/تغيير PIN).
- رسائل خطأ وتنبيهات واضحة في كل مسار.
"""
from functools import wraps

import json
import logging
import os

from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Avg, Count, Q, Subquery, OuterRef, FloatField, IntegerField
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from core.models import Profile, SchoolSettings
from students.models import ClassRoom, Student
from teachers.models import (
    ClassSession,
    ExamResult,
    StudentAnswer,
    Teacher,
    TeacherExam,
    TeacherQuestion,
    TeacherSkill,
)


# ═══════════════════════════════════════════════════════════════
# تشخيص الصور (صفحة مؤقتة)
# ═══════════════════════════════════════════════════════════════

def debug_images(request):
    """صفحة تشخيص لفحص إعدادات الصور و Cloudinary"""
    from django.conf import settings as django_settings
    from django.http import HttpResponse
    import cloudinary
    import cloudinary.api

    html = '<html dir="rtl"><head><meta charset="utf-8"><title>تشخيص الصور</title>'
    html += '<style>body{font-family:sans-serif;padding:20px;background:#f5f5f5}table{border-collapse:collapse;margin:10px 0}td,th{border:1px solid #ccc;padding:8px 12px;text-align:right}.ok{color:green;font-weight:bold}.err{color:red;font-weight:bold}img{border:2px solid #ccc;margin:5px}</style>'
    html += '</head><body>'
    html += '<h1>🔍 تشخيص الصور و Cloudinary</h1>'

    # 1) الإعدادات
    html += '<h2>1. الإعدادات</h2><table>'
    html += f'<tr><td>DEFAULT_FILE_STORAGE</td><td>{getattr(django_settings, "DEFAULT_FILE_STORAGE", "غير محدد")}</td></tr>'
    html += f'<tr><td>MEDIA_URL</td><td>{django_settings.MEDIA_URL}</td></tr>'
    html += f'<tr><td>CLOUD_NAME</td><td>{django_settings.CLOUDINARY_STORAGE.get("CLOUD_NAME", "؟")}</td></tr>'
    html += f'<tr><td>API_KEY</td><td>{django_settings.CLOUDINARY_STORAGE.get("API_KEY", "؟")[:6]}...</td></tr>'
    html += f'<tr><td>DEBUG</td><td>{django_settings.DEBUG}</td></tr>'

    # ترتيب INSTALLED_APPS
    apps = django_settings.INSTALLED_APPS
    cs_idx = apps.index('cloudinary_storage') if 'cloudinary_storage' in apps else -1
    sf_idx = apps.index('django.contrib.staticfiles') if 'django.contrib.staticfiles' in apps else -1
    order_ok = cs_idx >= 0 and sf_idx >= 0 and cs_idx < sf_idx
    html += f'<tr><td>ترتيب cloudinary_storage</td><td class="{"ok" if order_ok else "err"}">{"✅ قبل staticfiles (صحيح)" if order_ok else "❌ بعد staticfiles (خطأ!)"} — موقعه: {cs_idx}, staticfiles: {sf_idx}</td></tr>'
    html += '</table>'

    # 2) اختبار اتصال Cloudinary
    html += '<h2>2. اختبار اتصال Cloudinary</h2>'
    try:
        result = cloudinary.api.ping()
        html += f'<p class="ok">✅ الاتصال ناجح! النتيجة: {result}</p>'
    except Exception as e:
        html += f'<p class="err">❌ فشل الاتصال: {e}</p>'

    # 3) اختبار رفع صورة تجريبية
    html += '<h2>3. اختبار رفع صورة تجريبية</h2>'
    try:
        import cloudinary.uploader
        # رفع صورة بسيطة من URL عام
        test_result = cloudinary.uploader.upload(
            "https://res.cloudinary.com/demo/image/upload/sample.jpg",
            public_id="test_amane_platform",
            overwrite=True,
            folder="test"
        )
        test_url = test_result.get('secure_url', '')
        html += f'<p class="ok">✅ الرفع نجح!</p>'
        html += f'<p>URL: <a href="{test_url}" target="_blank">{test_url}</a></p>'
        html += f'<img src="{test_url}" style="max-height:100px">'
        # حذف الصورة التجريبية
        cloudinary.uploader.destroy("test/test_amane_platform")
    except Exception as e:
        html += f'<p class="err">❌ فشل الرفع: {e}</p>'

    # 4) الشعارات
    html += '<h2>4. الشعارات في قاعدة البيانات</h2>'
    ss = SchoolSettings.objects.first()
    if ss:
        html += '<table>'
        from cloudinary.utils import cloudinary_url as _cu
        for label, field in [('الوزارة', ss.ministry_logo), ('المدرسة', ss.school_logo)]:
            val = str(field) if field else ''
            html += f'<tr><td>شعار {label} — القيمة الخام</td><td><code>{val or "فارغ"}</code></td></tr>'
            if val:
                try:
                    cu, _ = _cu(val)
                    html += f'<tr><td>شعار {label} — Cloudinary URL</td><td><a href="{cu}" target="_blank">{cu}</a></td></tr>'
                except Exception as e:
                    html += f'<tr><td>شعار {label} — خطأ</td><td class="err">{e}</td></tr>'
        html += '</table>'

        # عرض الصور
        for label, field in [('الوزارة', ss.ministry_logo), ('المدرسة', ss.school_logo)]:
            val = str(field) if field else ''
            if val:
                cu, _ = _cu(val)
                html += f'<h3>شعار {label}:</h3><img src="{cu}" style="max-height:200px">'
    else:
        html += '<p class="err">❌ لا يوجد سجل SchoolSettings في قاعدة البيانات</p>'

    # 5) صور الأسئلة
    html += '<h2>5. صور الأسئلة (أول 5)</h2>'
    questions_with_images = TeacherQuestion.objects.exclude(question_image='').exclude(question_image__isnull=True)[:5]
    if questions_with_images:
        for q in questions_with_images:
            try:
                val = str(q.question_image)
                img_url, _ = _cu(val) if val and not val.startswith('http') else (val, None)
                html += f'<p>سؤال #{q.id}: {val}<br>URL: <a href="{img_url}" target="_blank">{img_url}</a><br><img src="{img_url}" style="max-height:100px"></p>'
            except Exception as e:
                html += f'<p>سؤال #{q.id}: <span class="err">خطأ — {e}</span></p>'
    else:
        html += '<p>لا توجد أسئلة بصور</p>'

    # 6) محتوى المهارات (صور + PDF)
    html += '<h2>6. محتوى المهارات (أول 5)</h2>'
    from teachers.models import TeacherSkillContent
    contents = TeacherSkillContent.objects.all()[:5]
    if contents:
        for c in contents:
            html += f'<div style="border:1px solid #ddd;padding:10px;margin:5px;background:white">'
            html += f'<b>محتوى #{c.id} (مهارة: {c.skill_id})</b><br>'
            # صورة
            img_val = str(c.plain_image) if c.plain_image else ''
            html += f'plain_image raw = <code>{img_val or "فارغ"}</code><br>'
            if img_val and img_val != 'None':
                try:
                    iu, _ = _cu(img_val)
                    html += f'URL: <a href="{iu}" target="_blank">{iu}</a><br><img src="{iu}" style="max-height:80px"><br>'
                except Exception as e:
                    html += f'<span class="err">خطأ: {e}</span><br>'
            # PDF
            pdf_val = str(c.pdf_file) if c.pdf_file else ''
            html += f'pdf_file raw = <code>{pdf_val or "فارغ"}</code><br>'
            if pdf_val and pdf_val != 'None':
                try:
                    from cloudinary.utils import cloudinary_url as _cu2
                    pu, _ = _cu2(pdf_val, resource_type='raw')
                    html += f'PDF URL: <a href="{pu}" target="_blank">{pu}</a><br>'
                except Exception as e:
                    html += f'<span class="err">خطأ: {e}</span><br>'
            # فيديو
            html += f'video_url = <code>{c.video_url or "فارغ"}</code><br>'
            html += f'plain_text = <code>{(c.plain_text or "")[:100] or "فارغ"}</code>'
            html += '</div>'
    else:
        html += '<p>لا يوجد محتوى</p>'

    # 7) متغيرات البيئة المتعلقة بـ Cloudinary
    html += '<h2>7. متغيرات البيئة</h2><table>'
    for env_key in ['CLOUDINARY_URL', 'CLOUDINARY_CLOUD_NAME', 'CLOUDINARY_API_KEY']:
        env_val = os.environ.get(env_key, '')
        masked = env_val[:10] + '...' if env_val else 'غير موجود'
        html += f'<tr><td>{env_key}</td><td><code>{masked}</code></td></tr>'
    html += '</table>'

    html += '<hr><p style="color:#999">بعد التأكد من عمل الصور، احذفي هذا المسار من urls.py</p>'
    html += '</body></html>'
    return HttpResponse(html)


# ═══════════════════════════════════════════════════════════════
# Decorators وأدوات صلاحيات
# ═══════════════════════════════════════════════════════════════

def _get_role(user):
    """يرجع دور المستخدم (ADMIN/TEACHER/STUDENT) أو None."""
    if not user.is_authenticated:
        return None
    try:
        return user.core_profile.role
    except Profile.DoesNotExist:
        return None


def admin_required(view_func):
    """يضمن أن المستخدم مسجل دخول ودوره ADMIN (أو superuser)."""
    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        if request.user.is_superuser or _get_role(request.user) == 'ADMIN':
            return view_func(request, *args, **kwargs)
        messages.error(request, 'هذه الصفحة مخصصة للإدارة فقط.')
        return redirect('home')
    return _wrapped


def _safe_int(value, default, minimum=None, maximum=None):
    """تحويل آمن للأعداد مع حدود اختيارية — لا يطرح ValueError."""
    try:
        result = int(str(value).strip())
    except (TypeError, ValueError):
        return default
    if minimum is not None and result < minimum:
        return minimum
    if maximum is not None and result > maximum:
        return maximum
    return result


# جداول تحويل الأرقام (لا نُغيّر المُخزَّن — نحوّل فقط للمقارنة):
# الأرقام العربية ↔ اللاتينية، لتمكين البحث المتساهل بصيغتين.
_AR_TO_LA = str.maketrans({
    '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
    '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9',
    '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
    '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9',
})
_LA_TO_AR = str.maketrans({
    '0': '٠', '1': '١', '2': '٢', '3': '٣', '4': '٤',
    '5': '٥', '6': '٦', '7': '٧', '8': '٨', '9': '٩',
})


def _id_variants(value):
    """
    يرجع كل صيغ رقم الهوية للبحث:
    1) الصيغة كما هي (كما كتبتْها المستخدمة).
    2) صيغة لاتينية (لمواجهة بيانات قديمة).
    3) صيغة عربية (لمواجهة لوحات مفاتيح iPhone).
    """
    if not value:
        return []
    s = str(value).strip()
    return list({s, s.translate(_AR_TO_LA), s.translate(_LA_TO_AR)})


def _is_valid_id(value):
    """تحقق من رقم الهوية بعد تطبيعه: 8-12 رقماً عددياً."""
    if not value:
        return False
    latin = str(value).translate(_AR_TO_LA).strip()
    return latin.isdigit() and 8 <= len(latin) <= 12


# ═══════════════════════════════════════════════════════════════
# الصفحات العامة + صلاحيات المحتوى للطالبة
# ═══════════════════════════════════════════════════════════════

# أنواع المحتوى التي يحق للطالبة رؤيتها في الصفحة الرئيسية:
#   - skill        : مهارة قدرات (تحوي محتوى + قبلي/بعدي، لكنها تُعرض هنا
#                    كبطاقة محتوى. فلترة الاختبار القبلي/البعدي تتم في
#                    skill_detail و student dashboard.)
#   - lesson       : درس تحصيلي (محتوى فقط للطالبة).
#   - comprehensive: اختبار شامل (فقط إن كان مفعّلاً).
#
# يُستثنى من الطالبة: bank (بنك أسئلة) — وكذلك الاختبارات بأنواعها.
STUDENT_VISIBLE_CONTENT = {'skill', 'lesson', 'comprehensive'}


def home(request):
    """
    الصفحة الرئيسية — تعرض المحتوى التعليمي + الاختبارات الشاملة.

    منطق العرض حسب الدور:
    - زائرة / طالبة → فقط: مهارات قدرات + دروس تحصيلي + شاملة مفعّلة.
      (لا تظهر بنوك الأسئلة ولا اختبارات التحصيلي).
    - معلمة / مديرة → كل المحتوى المفعّل + الشاملة (للإدارة).
    """
    settings_obj = SchoolSettings.objects.first()
    role = _get_role(request.user)

    base_qs = (
        TeacherSkill.objects
        .select_related('created_by')
        .order_by('-created_at')
    )

    if role in ('ADMIN', 'TEACHER'):
        # كل المحتوى الفعّال + الشاملة دائماً
        skills = base_qs.filter(
            Q(is_active=True) | Q(content_type='comprehensive')
        )
    else:
        # طالبة/زائرة: فلترة صارمة + إخفاء بنوك الأسئلة وأي اختبار غير شامل
        skills = base_qs.filter(
            content_type__in=STUDENT_VISIBLE_CONTENT,
        ).filter(
            Q(is_active=True) | Q(content_type='comprehensive', is_active=True)
        )

    return render(request, 'core/home.html', {
        'settings': settings_obj,
        'skills': skills,
        'classrooms': ClassRoom.objects.all(),
        'role': role,
    })


def skill_detail(request, skill_id):
    """
    تفاصيل المهارة/الدرس + الاختبارات.

    قيود الطالبة:
    - لا يحق لها رؤية اختبار قبلي/بعدي/درس/بنك إلا عبر "أداء الاختبار"
      المفعّل من المعلمة (تأخذ الرابط منها). هنا نعرض المحتوى التعليمي فقط.
    """
    skill = get_object_or_404(TeacherSkill, pk=skill_id)
    role = _get_role(request.user)

    exams = list(skill.exams.all())
    pre_exam = next((e for e in exams if e.exam_type == 'pre'), None)
    post_exam = next((e for e in exams if e.exam_type == 'post'), None)
    lesson_exam = next((e for e in exams if e.exam_type == 'lesson'), None)
    bank_exam = next((e for e in exams if e.exam_type == 'bank'), None)
    comp_exam = next(
        (e for e in exams if e.exam_type in ('comprehensive_qodrat', 'comprehensive_tahsili')),
        None,
    )

    # للطالبة/الزائرة: نخفي تفاصيل الاختبارات تماماً
    is_student = (role == 'STUDENT') or role is None
    if is_student:
        pre_exam = None
        post_exam = None
        lesson_exam = None
        bank_exam = None
        # نسمح برؤية الاختبار الشامل فقط إن كان مفعّلاً
        if comp_exam and not comp_exam.is_active:
            comp_exam = None

    return render(request, 'core/skill_detail.html', {
        'skill': skill,
        'pre_exam': pre_exam,
        'post_exam': post_exam,
        'lesson_exam': lesson_exam,
        'bank_exam': bank_exam,
        'comp_exam': comp_exam,
        'classrooms': ClassRoom.objects.all(),
        'role': role,
        'is_student': is_student,
    })


def take_test(request, skill_id):
    """عرض الاختبار القبلي العام (للزائر)."""
    skill = get_object_or_404(TeacherSkill, pk=skill_id)
    pre_exam = skill.exams.filter(exam_type='pre').first()
    questions = pre_exam.questions.all() if pre_exam else []
    return render(request, 'core/take_test.html', {
        'skill': skill,
        'questions': questions,
    })


def skill_print(request, skill_id):
    """نسخة قابلة للطباعة من المحتوى التعليمي للمهارة/الدرس."""
    skill = get_object_or_404(TeacherSkill, pk=skill_id)
    content = getattr(skill, 'content', None)
    return render(request, 'core/skill_print.html', {
        'skill': skill,
        'content': content,
        'settings': SchoolSettings.objects.first(),
    })


@login_required
def exam_print(request, exam_id):
    """
    نسخة قابلة للطباعة من الاختبار (مع/بدون الإجابات).
    - المعلمة/المديرة: تظهر الإجابات الصحيحة (نسخة المعلم).
    - الطالبة: نموذج فارغ بدون إجابات صحيحة.
    """
    exam = get_object_or_404(TeacherExam, pk=exam_id)
    role = _get_role(request.user)
    show_answers = role in ('ADMIN', 'TEACHER') and request.GET.get('answers') != '0'
    return render(request, 'core/exam_print.html', {
        'exam': exam,
        'questions': exam.questions.all().order_by('order'),
        'show_answers': show_answers,
        'settings': SchoolSettings.objects.first(),
    })


# ═══════════════════════════════════════════════════════════════
# تفعيل/إلغاء الاختبارات (AJAX)
# ═══════════════════════════════════════════════════════════════

@login_required
@require_POST
def activate_exam(request, exam_id):
    """
    تفعيل الاختبار لفصل محدد، أو لجميع الفصول إن كان classroom_id == 'all'.
    يولّد رابطاً مخصصاً + يسجّل حصة (للمعلمات).
    """
    exam = get_object_or_404(TeacherExam, pk=exam_id)
    classroom_id = (request.POST.get('classroom_id') or '').strip()
    if not classroom_id:
        return JsonResponse({'success': False, 'error': 'لم يُحدّد فصل'}, status=400)

    # تفعيل الاختبار نفسه
    exam.is_active = True
    exam.save(update_fields=['is_active'])

    # تخزين الفصول المستهدفة على المهارة الأم لتُعرض على الطالبة
    if classroom_id == 'all':
        target_text = 'جميع الفصول'
    else:
        cls = get_object_or_404(ClassRoom, pk=classroom_id)
        target_text = cls.name

    skill = exam.skill
    # نضمّن في target_classes (string) — توافقي مع باقي الكود
    skill.target_classes = target_text
    skill.is_active = True  # تفعيل المهارة لتظهر في الواجهة
    skill.save(update_fields=['target_classes', 'is_active'])

    # تسجيل حصة إذا كان من فعّل معلمة
    try:
        teacher = Teacher.objects.get(user=request.user)
        ClassSession.objects.get_or_create(
            teacher=teacher,
            skill=skill,
            session_type='qodrat' if skill.skill_type in ('qodrat_kamy', 'qodrat_lafzy') else 'tahsili',
            target_class=target_text,
            session_date=timezone.now().date(),
            defaults={
                'session_time': timezone.now().time(),
                'exam': exam,
            },
        )
    except Teacher.DoesNotExist:
        pass

    url = request.build_absolute_uri(f'/students/exam/{exam.id}/')
    return JsonResponse({
        'success': True,
        'url': url,
        'target': target_text,
        'message': f'✅ تم تفعيل {exam.get_exam_type_display()} — {target_text}',
    })


@login_required
@require_POST
def deactivate_exam(request, exam_id):
    """إلغاء تفعيل اختبار."""
    exam = get_object_or_404(TeacherExam, pk=exam_id)
    exam.is_active = False
    exam.save(update_fields=['is_active'])
    return JsonResponse({'success': True})


# ═══════════════════════════════════════════════════════════════
# لوحة المديرة
# ═══════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════
# Admin V2 — Enterprise Dashboard (Single-Page)
# ═══════════════════════════════════════════════════════════════

def _v2_kpis():
    """KPIs الرئيسية + مؤشر التحسن (مقارنة بالـ 30 يوم السابقة)."""
    from datetime import timedelta
    now = timezone.now()
    last_30 = now - timedelta(days=30)
    prev_30 = now - timedelta(days=60)

    total_students = Profile.objects.filter(role='STUDENT').count()
    total_teachers = Profile.objects.filter(role='TEACHER').count()
    active_exams = TeacherExam.objects.filter(is_active=True).count()
    active_sessions = ClassSession.objects.filter(session_date__gte=last_30.date()).count()

    avg_now = ExamResult.objects.filter(
        submitted_at__gte=last_30
    ).aggregate(a=Avg('percentage'))['a'] or 0
    avg_prev = ExamResult.objects.filter(
        submitted_at__gte=prev_30, submitted_at__lt=last_30
    ).aggregate(a=Avg('percentage'))['a'] or 0
    delta = round(avg_now - avg_prev, 1)

    avg_total = ExamResult.objects.aggregate(a=Avg('percentage'))['a'] or 0

    return {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'active_exams': active_exams,
        'active_sessions': active_sessions,
        'avg_score': round(avg_total, 1),
        'avg_score_30d': round(avg_now, 1),
        'improvement_delta': delta,
        'improvement_pct': round((delta / max(avg_prev, 1)) * 100, 1) if avg_prev else 0,
    }


def _v2_time_series(days=30):
    """سلسلة زمنية لـ Line chart — متوسط النسبة + عدد المحاولات يومياً."""
    from datetime import timedelta
    now = timezone.now().date()
    series = []
    for i in range(days - 1, -1, -1):
        d = now - timedelta(days=i)
        day_results = ExamResult.objects.filter(submitted_at__date=d)
        series.append({
            'date': d.isoformat(),
            'count': day_results.count(),
            'avg': round(day_results.aggregate(a=Avg('percentage'))['a'] or 0, 1),
        })
    return series


def _v2_classroom_compare():
    """مقارنة أداء الفصول — Bar chart."""
    rooms = []
    for cls in ClassRoom.objects.all().order_by('name'):
        names = list(cls.students.values_list('full_name', flat=True))
        if not names:
            rooms.append({'name': cls.name, 'avg': 0, 'count': 0, 'students': 0})
            continue
        # نطابق بالاسم الكامل (تصميم البيانات الحالي)
        results = ExamResult.objects.filter(
            student__in=User.objects.filter(
                Q(first_name__in=[n.split()[0] for n in names if n]) |
                Q(username__in=names)
            )
        )
        avg = results.aggregate(a=Avg('percentage'))['a'] or 0
        rooms.append({
            'name': cls.name,
            'avg': round(avg, 1),
            'count': results.count(),
            'students': len(names),
        })
    return rooms


def _v2_level_distribution():
    """توزيع الطالبات على مستويات الأداء — Pie/Donut. الطالبات فقط."""
    perf = (
        ExamResult.objects
        .filter(student__core_profile__role='STUDENT')
        .exclude(student__is_superuser=True)
        .values('student_id')
        .annotate(avg=Avg('percentage'))
    )
    excellent = mid = weak = unknown = 0
    for row in perf:
        a = row['avg'] or 0
        if a >= 90:
            excellent += 1
        elif a >= 50:
            mid += 1
        else:
            weak += 1
    total_students = Profile.objects.filter(role='STUDENT').count()
    rated = excellent + mid + weak
    unknown = max(0, total_students - rated)
    return {
        'excellent': excellent, 'mid': mid, 'weak': weak, 'untested': unknown,
    }


def _v2_top_students(limit=10):
    """أعلى الطالبات أداءً — Leaderboard. الطالبات فقط — لا مديرة ولا معلمات."""
    perf = (
        ExamResult.objects
        # نستبعد أي User دوره ليس STUDENT (المديرة، المعلمات، superusers)
        .filter(student__core_profile__role='STUDENT')
        .exclude(student__is_superuser=True)
        .values('student_id', 'student__first_name', 'student__last_name', 'student__username')
        .annotate(avg=Avg('percentage'), count=Count('id'))
        .order_by('-avg', '-count')[:limit]
    )
    out = []
    for r in perf:
        name = (f"{r['student__first_name']} {r['student__last_name']}".strip()
                or r['student__username'])
        out.append({
            'name': name,
            'avg': round(r['avg'] or 0, 1),
            'count': r['count'],
            'student_id': r['student_id'],
        })
    return out


def _v2_teachers_impact():
    """
    تحليل أداء المعلمات — أرقام دقيقة بدون تضخم.

    الإصلاحات:
    - كل مقياس يُحسب بـ Subquery منفصل → لا تضخم من JOINs متداخلة.
    - التأثير الحقيقي = متوسط البعدي − متوسط القبلي (تحسّن فعلي).
      · لو عندها بعدي فقط → متوسط البعدي.
      · لو لا يوجد بعدي ولا قبلي → المتوسط العام.
    """
    # ── Subqueries منفصلة لكل مقياس ─────────────────────────────
    def _count_sq(qs, group_field):
        return Subquery(
            qs.filter(**{group_field: OuterRef('pk')})
              .values(group_field)
              .annotate(c=Count('id'))
              .values('c')[:1],
            output_field=IntegerField()
        )

    def _avg_sq(exam_type_filter=None):
        qs = ExamResult.objects.filter(exam__skill__created_by=OuterRef('pk'))
        if exam_type_filter:
            qs = qs.filter(exam__exam_type=exam_type_filter)
        return Subquery(
            qs.values('exam__skill__created_by')
              .annotate(avg=Avg('percentage'))
              .values('avg')[:1],
            output_field=FloatField()
        )

    teachers = (
        Teacher.objects
        .select_related('user')
        .annotate(
            sessions_count=_count_sq(ClassSession.objects.all(), 'teacher'),
            skills_count=_count_sq(TeacherSkill.objects.all(), 'created_by'),
            exams_count=Subquery(
                TeacherExam.objects
                .filter(skill__created_by=OuterRef('pk'))
                .values('skill__created_by')
                .annotate(c=Count('id'))
                .values('c')[:1],
                output_field=IntegerField()
            ),
            results_count=Subquery(
                ExamResult.objects
                .filter(exam__skill__created_by=OuterRef('pk'))
                .values('exam__skill__created_by')
                .annotate(c=Count('id'))
                .values('c')[:1],
                output_field=IntegerField()
            ),
            pre_avg=_avg_sq('pre'),
            post_avg=_avg_sq('post'),
            all_avg=_avg_sq(),
        )
    )

    out = []
    for t in teachers:
        pre  = float(t.pre_avg  or 0)
        post = float(t.post_avg or 0)
        avg  = float(t.all_avg  or 0)

        # التأثير الحقيقي: تحسن من القبلي للبعدي
        if pre > 0 and post > 0:
            # نسبة التحسن (0-100 scale: نرفعها لتكون قابلة للمقارنة)
            improvement = round(post - pre, 1)
            impact = round(post, 1)       # للعرض: متوسط البعدي
            has_improvement = True
        elif post > 0:
            improvement = None
            impact = round(post, 1)
            has_improvement = False
        else:
            improvement = None
            impact = round(avg, 1)
            has_improvement = False

        out.append({
            'id':          t.id,
            'name':        t.full_name,
            'sessions':    int(t.sessions_count  or 0),
            'skills':      int(t.skills_count    or 0),
            'exams':       int(t.exams_count     or 0),
            'engagement':  int(t.results_count   or 0),
            'impact':      impact,            # للألوان والعرض (متوسط البعدي أو العام)
            'pre_avg':     round(pre,  1),
            'post_avg':    round(post, 1),
            'improvement': improvement,       # None لو ما عندها قبلي+بعدي
            'has_improvement': has_improvement,
        })

    out.sort(key=lambda x: -(x['improvement'] or x['impact']))
    return out


def _v2_active_exams_inline():
    """قائمة الاختبارات الفعّالة — للإدارة المضمّنة."""
    exams = (
        TeacherExam.objects
        .select_related('skill', 'skill__created_by')
        .annotate(participants=Count('results', distinct=True))
        .order_by('-is_active', '-created_at')[:50]
    )
    out = []
    for e in exams:
        out.append({
            'id': e.id,
            'skill_id': e.skill_id,
            'title': e.skill.title,
            'type': e.get_exam_type_display(),
            'is_active': e.is_active,
            'participants': e.participants or 0,
            'duration': e.duration_minutes,
            'pass_score': e.pass_score,
            'teacher': e.skill.created_by.full_name if e.skill.created_by else '—',
            'is_comprehensive': e.skill.content_type == 'comprehensive',
        })
    return out


def _v2_smart_alerts(kpis, level_dist, teachers_impact, top_students):
    """تنبيهات ذكية + استنتاجات + توصيات."""
    alerts = []

    # 1) ضعف الأداء العام
    if kpis['avg_score'] and kpis['avg_score'] < 60:
        alerts.append({
            'level': 'danger',
            'icon': '⚠️',
            'title': f"متوسط الفصل {kpis['avg_score']}% — أقل من حد الأداء المطلوب",
            'msg': 'يُنصح بمراجعة المهارات الأكثر صعوبة على الطالبات وإعادة شرحها.',
        })

    # 2) تحسن أو تدهور
    if kpis['improvement_delta'] >= 5:
        alerts.append({
            'level': 'success',
            'icon': '🎉',
            'title': f"الأداء يتحسن — +{kpis['improvement_delta']}% خلال الـ 30 يوم الأخيرة",
            'msg': 'استمري في نفس النهج التعليمي — البيانات تُظهر تأثيراً إيجابياً.',
        })
    elif kpis['improvement_delta'] <= -5:
        alerts.append({
            'level': 'warning',
            'icon': '📉',
            'title': f"الأداء يتراجع — {kpis['improvement_delta']}% خلال الـ 30 يوم الأخيرة",
            'msg': 'راجعي المعلمات وحاولي تحديد سبب التراجع (مهارة معيّنة؟ فصل معيّن؟).',
        })

    # 3) عدد كبير في فئة الضعف
    if level_dist['weak'] > level_dist['excellent'] and level_dist['weak'] >= 3:
        alerts.append({
            'level': 'warning',
            'icon': '🆘',
            'title': f"{level_dist['weak']} طالبة تحت 50% — يحتجن دعماً عاجلاً",
            'msg': 'فعّلي خطة دعم فردية أو حصصاً علاجية لهؤلاء الطالبات.',
        })

    # 4) معلمة تحتاج دعم
    if teachers_impact:
        worst = min((t for t in teachers_impact if t['impact'] > 0),
                    key=lambda t: t['impact'], default=None)
        best = max(teachers_impact, key=lambda t: t['impact'], default=None)
        if worst and best and best['impact'] - worst['impact'] >= 20:
            alerts.append({
                'level': 'info',
                'icon': '👩‍🏫',
                'title': f"{best['name']} الأعلى تأثيراً ({best['impact']}%) • {worst['name']} الأقل ({worst['impact']}%)",
                'msg': f'يمكن لـ {best["name"]} مشاركة خبرتها مع {worst["name"]} لدعم الأداء.',
            })

    # 5) تشجيع الطالبات المتفوقات
    if top_students and top_students[0]['avg'] >= 90:
        alerts.append({
            'level': 'success',
            'icon': '🏆',
            'title': f"تتصدر {top_students[0]['name']} الفصل بـ {top_students[0]['avg']}%",
            'msg': 'فكّري في تكريمها أمام الفصل لتشجيع الطالبات الأخريات.',
        })

    return alerts


def _v2_decisions(kpis, level_dist, alerts):
    """مركز القرار — 3 إجراءات فورية مبنية على البيانات."""
    decisions = []
    if level_dist['weak'] >= 3:
        decisions.append({
            'icon': '🎯',
            'priority': 'عاجل',
            'action': 'إعداد جلسة دعم للطالبات الضعيفات',
            'why': f"{level_dist['weak']} طالبة تحت 50% — تأخير المعالجة يوسّع الفجوة.",
        })
    if kpis['active_exams'] == 0:
        decisions.append({
            'icon': '📝',
            'priority': 'مهم',
            'action': 'تفعيل اختبار قياس جديد',
            'why': 'لا توجد اختبارات نشطة حالياً — تأكدي من جداول المعلمات.',
        })
    if kpis['improvement_delta'] <= -5:
        decisions.append({
            'icon': '📊',
            'priority': 'مهم',
            'action': 'تحليل سبب تراجع الأداء',
            'why': f"تراجع {kpis['improvement_delta']}% — راجعي تقارير المهارات والفصول.",
        })
    if not decisions:
        decisions.append({
            'icon': '✨',
            'priority': 'تطوير',
            'action': 'إضافة محتوى تدريبي إضافي',
            'why': 'الأداء مستقر — هذا وقت ممتاز للتوسّع وإثراء المنصة.',
        })
    if level_dist['excellent'] >= 3:
        decisions.append({
            'icon': '🏆',
            'priority': 'تحفيز',
            'action': 'تكريم المتفوقات',
            'why': f"{level_dist['excellent']} طالبة فوق 90% — التحفيز يبني ثقافة التميّز.",
        })
    return decisions[:3]


def _v2_predictions():
    """تنبؤ بسيط — يعتمد على متوسط آخر 7 أيام لتقدير المتوقع."""
    from datetime import timedelta
    now = timezone.now()
    last_week = ExamResult.objects.filter(
        submitted_at__gte=now - timedelta(days=7)
    ).aggregate(a=Avg('percentage'))['a'] or 0
    last_month = ExamResult.objects.filter(
        submitted_at__gte=now - timedelta(days=30)
    ).aggregate(a=Avg('percentage'))['a'] or 0
    momentum = last_week - last_month
    forecast = max(0, min(100, last_week + momentum * 0.5))
    return {
        'last_week': round(last_week, 1),
        'last_month': round(last_month, 1),
        'forecast': round(forecast, 1),
        'momentum': round(momentum, 1),
    }


@admin_required
def admin_v2_dashboard(request):
    """لوحة المديرة الجديدة — Single-Page Enterprise Dashboard."""
    kpis = _v2_kpis()
    time_series = _v2_time_series()
    classroom_cmp = _v2_classroom_compare()
    level_dist = _v2_level_distribution()
    top_students = _v2_top_students(limit=10)
    teachers_impact = _v2_teachers_impact()
    active_exams = _v2_active_exams_inline()
    alerts = _v2_smart_alerts(kpis, level_dist, teachers_impact, top_students)
    decisions = _v2_decisions(kpis, level_dist, alerts)
    predictions = _v2_predictions()

    # قوائم لإدارة المستخدمين (modals)
    teacher_profiles = (
        Profile.objects.filter(role='TEACHER').select_related('user')
        .annotate(
            results_count=Count('user__teacher_exam_results', distinct=True),
            avg_score=Avg('user__teacher__teacher_skills__exams__results__percentage'),
        )
    )
    student_profiles = (
        Profile.objects.filter(role='STUDENT').select_related('user')
        .annotate(
            results_count=Count('user__teacher_exam_results', distinct=True),
            avg_score=Avg('user__teacher_exam_results__percentage'),
        )
    )

    teachers_list = [{
        'id': p.user.id, 'name': p.user.get_full_name() or p.user.username,
        'national_id': p.national_id, 'pin': p.pin_code,
        'is_active': p.user.is_active,
        'results': p.results_count or 0,
        'avg': round(p.avg_score or 0, 1),
        'teacher_classrooms': list(p.classrooms.values_list('id', flat=True)) if hasattr(p, 'classrooms') else [],
    } for p in teacher_profiles]

    students_list = []
    student_index = {}
    for s in Student.objects.select_related('classroom').all():
        student_index.setdefault(s.full_name, s)
    for p in student_profiles:
        full_name = p.user.get_full_name() or p.user.username
        s = student_index.get(full_name)
        students_list.append({
            'id': p.user.id, 'name': full_name,
            'national_id': p.national_id,
            'classroom': s.classroom.name if s and s.classroom else '—',
            'is_active': p.user.is_active,
            'results': p.results_count or 0,
            'avg': round(p.avg_score or 0, 1),
        })

    comprehensive_skills = (
        TeacherSkill.objects.filter(content_type='comprehensive')
        .select_related('created_by').prefetch_related('exams').order_by('-created_at')
    )

    # وضع العرض المدرسي
    display_mode = request.GET.get('display') == '1'

    return render(request, 'core/admin_v2.html', {
        'kpis': kpis,
        'time_series_json': _to_json(time_series),
        'classrooms_json': _to_json(classroom_cmp),
        'levels_json': _to_json(level_dist),
        'top_students': top_students,
        'teachers_impact': teachers_impact,
        'active_exams': active_exams,
        'alerts': alerts,
        'decisions': decisions,
        'predictions': predictions,
        'teachers_list': teachers_list,
        'students_list': students_list,
        'comprehensive_skills': comprehensive_skills,
        'classrooms': ClassRoom.objects.all().order_by('name'),
        'display_mode': display_mode,
    })


def _to_json(data):
    """تحويل آمن لـ JSON ضمن قالب."""
    import json
    return json.dumps(data, default=str, ensure_ascii=False)


@admin_required
def admin_v2_data_json(request):
    """API للتحديثات الحية — يرجع JSON نظيف."""
    from django.http import JsonResponse
    return JsonResponse({
        'kpis': _v2_kpis(),
        'time_series': _v2_time_series(),
        'classrooms': _v2_classroom_compare(),
        'levels': _v2_level_distribution(),
        'top_students': _v2_top_students(),
        'teachers_impact': _v2_teachers_impact(),
        'predictions': _v2_predictions(),
    })


@admin_required
def admin_assign_teacher_classrooms(request, teacher_id):
    """تعيين / إلغاء تعيين الفصول لمعلمة — يقبل POST فقط."""
    from django.http import JsonResponse
    from teachers.models import Teacher as TeacherProfile
    if request.method != 'POST':
        return JsonResponse({'error': 'method not allowed'}, status=405)
    try:
        teacher_profile = TeacherProfile.objects.get(user_id=teacher_id)
    except TeacherProfile.DoesNotExist:
        return JsonResponse({'error': 'teacher not found'}, status=404)
    classroom_ids = request.POST.getlist('classroom_ids[]')
    teacher_profile.classrooms.set(classroom_ids)
    assigned = list(teacher_profile.classrooms.values('id', 'name'))
    return JsonResponse({'ok': True, 'assigned': assigned})


# ═══════════════════════════════════════════════════════════════
# Advanced Analytics + Reports
# ═══════════════════════════════════════════════════════════════

def _filter_results(scope, target_id, exam_type):
    """يبني QuerySet للنتائج حسب الفلتر."""
    qs = ExamResult.objects.select_related('exam', 'exam__skill', 'exam__skill__created_by', 'student')

    if scope == 'classroom' and target_id:
        # نطابق بالاسم الكامل (التصميم الحالي)
        names = list(Student.objects.filter(classroom__name=target_id).values_list('full_name', flat=True))
        if not names:
            return qs.none()
        # ربط بالـ User: نقابل بـ first_name و username
        from django.db.models import Q
        cond = Q()
        valid_cond = False
        for n in names:
            parts = n.split()
            if parts and parts[0]:
                cond |= Q(student__first_name=parts[0]) | Q(student__username=n)
                valid_cond = True
        if not valid_cond:
            return qs.none()
        qs = qs.filter(cond)
    elif scope == 'teacher' and target_id:
        try:
            qs = qs.filter(exam__skill__created_by_id=int(target_id))
        except ValueError:
            return qs.none()
    elif scope == 'student' and target_id:
        try:
            qs = qs.filter(student_id=int(target_id))
        except ValueError:
            return qs.none()

    if exam_type and exam_type != 'all':
        qs = qs.filter(exam__exam_type=exam_type)

    return qs


def _descriptive_text(scope, avg, attempts, pass_pct, improvement):
    """يولّد جملة وصفية ذكية حسب الأداء."""
    if attempts == 0:
        return '⚠️ لا توجد محاولات لهذه الفئة في الفترة المحددة. ابدئي بتفعيل الاختبارات.'

    descriptors = []
    # المستوى العام
    if avg >= 90:
        descriptors.append('🏆 <strong style="color:var(--green)">أداء ممتاز</strong>')
    elif avg >= 75:
        descriptors.append('🌟 <strong style="color:var(--blue)">أداء جيد جداً</strong>')
    elif avg >= 60:
        descriptors.append('👍 <strong style="color:var(--orange)">أداء جيد</strong>')
    elif avg >= 50:
        descriptors.append('⚠️ <strong style="color:var(--orange)">أداء مقبول — يحتاج تطوير</strong>')
    else:
        descriptors.append('🆘 <strong style="color:var(--red)">أداء ضعيف — يحتاج تدخل عاجل</strong>')

    # السلوك
    if attempts >= 10:
        descriptors.append('🔥 <strong>نشطة جداً</strong> في خوض الاختبارات (' + str(attempts) + ' محاولة)')
    elif attempts >= 5:
        descriptors.append('✅ <strong>تفاعل جيد</strong> (' + str(attempts) + ' محاولة)')
    else:
        descriptors.append('⚠️ <strong>تفاعل محدود</strong> (' + str(attempts) + ' محاولة فقط) — حفّزيها على المحاولة')

    # التحسن
    if improvement >= 10:
        descriptors.append('📈 <strong style="color:var(--green)">تحسّن ملموس</strong> بمقدار +' + str(improvement) + '%')
    elif improvement <= -10:
        descriptors.append('📉 <strong style="color:var(--red)">تراجع</strong> بمقدار ' + str(improvement) + '% — راجعي السبب')

    # نسبة النجاح
    if pass_pct >= 80:
        descriptors.append('🎯 نسبة نجاح عالية: ' + str(pass_pct) + '%')
    elif pass_pct < 50:
        descriptors.append('⚠️ نسبة الرسوب مرتفعة: ' + str(100 - pass_pct) + '%')

    intro_map = {
        'school': 'تحليل أداء المدرسة كاملة:',
        'classroom': 'تحليل أداء الفصل:',
        'teacher': 'تحليل أداء طالبات هذه المعلمة:',
        'student': 'تحليل أداء الطالبة:',
    }
    intro = intro_map.get(scope, 'تحليل الأداء:')
    return intro + '<br>' + '<br>'.join(descriptors)


def _skills_breakdown(qs):
    """تحليل أداء المهارات — الأسئلة المرتبطة بـ target_skill_name."""
    from collections import defaultdict
    bucket = defaultdict(lambda: {'correct': 0, 'total': 0})

    for ans in StudentAnswer.objects.filter(result__in=qs).select_related('question'):
        name = (ans.question.target_skill_name or 'غير محدّدة').strip() or 'غير محدّدة'
        bucket[name]['total'] += 1
        if ans.is_correct:
            bucket[name]['correct'] += 1

    rows = []
    mastered = 0
    needs = 0
    for name, v in bucket.items():
        if v['total'] == 0:
            continue
        pct = round(100 * v['correct'] / v['total'])
        rows.append({'name': name, 'pct': pct, 'correct': v['correct'], 'total': v['total']})
        if pct >= 70:
            mastered += 1
        else:
            needs += 1
    rows.sort(key=lambda r: -r['pct'])
    return rows[:12], mastered, needs


def _exam_types_breakdown(qs):
    """متوسط الأداء حسب نوع الاختبار."""
    from collections import defaultdict
    bucket = defaultdict(lambda: {'sum': 0, 'count': 0, 'label': ''})
    type_labels = {
        'pre': 'قبلي',
        'post': 'بعدي',
        'lesson': 'درس تحصيلي',
        'bank': 'بنك أسئلة',
        'comprehensive_qodrat': 'شامل قدرات',
        'comprehensive_tahsili': 'شامل تحصيلي',
    }
    for r in qs:
        t = r.exam.exam_type
        bucket[t]['sum'] += float(r.percentage or 0)
        bucket[t]['count'] += 1
        bucket[t]['label'] = type_labels.get(t, t)

    out = []
    for k, v in bucket.items():
        if v['count']:
            out.append({'type': k, 'label': v['label'], 'avg': round(v['sum'] / v['count'], 1), 'count': v['count']})
    out.sort(key=lambda x: -x['avg'])
    return out


@admin_required
def admin_analytics_json(request):
    """يرجع التحليل المتقدم بناءً على الفلاتر."""
    from django.http import JsonResponse
    from datetime import timedelta

    try:
        return _admin_analytics_inner(request)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'avg': None, 'attempts': 0, 'pass_pct': None, 'fail_pct': None,
            'highest': None, 'lowest': None, 'improvement': None,
            'descriptive': f'❌ خطأ في التحليل: {exc}',
            'skills_mastered': 0, 'skills_needs': 0, 'skills': [], 'types': [],
            'no_data': True,
        }, status=200)


def _admin_analytics_inner(request):
    from django.http import JsonResponse

    scope = request.GET.get('scope', 'school')
    target_id = request.GET.get('id', '')
    exam_type = request.GET.get('exam_type', 'all')

    qs = _filter_results(scope, target_id, exam_type)
    total_qs = qs.count()
    if total_qs == 0:
        # نتحقق إن كانت توجد بيانات للفئة بأي نوع آخر — لمساعدة المعلمة على التشخيص
        any_qs = _filter_results(scope, target_id, 'all')
        any_count = any_qs.count()
        available_types = list(any_qs.values_list('exam__exam_type', flat=True).distinct()) if any_count else []
        type_labels = {
            'pre': 'قبلي', 'post': 'بعدي', 'lesson': 'درس تحصيلي',
            'bank': 'بنك أسئلة', 'comprehensive_qodrat': 'شامل قدرات',
            'comprehensive_tahsili': 'شامل تحصيلي',
        }
        type_label = type_labels.get(exam_type, exam_type)

        if any_count == 0:
            descriptive = (
                '<strong>📭 لا توجد محاولات اختبارات لهذه الفئة بعد.</strong><br>'
                '<span style="font-size:11px;color:var(--text3)">الأسباب المحتملة:</span><br>'
                '• الطالبة/الفصل/المعلمة لم تبدأ خوض أي اختبار حتى الآن.<br>'
                '• لم يُفعَّل أي اختبار لهذه الفئة من قبل المعلمات.<br>'
                '<br><span style="color:var(--blue)">💡 الحل: فعّلي اختباراً جديداً أو انتظري حتى تخوض الطالبات الاختبارات.</span>'
            )
        else:
            available_str = '، '.join(type_labels.get(t, t) for t in available_types) if available_types else '—'
            descriptive = (
                f'<strong>⚠️ لا توجد محاولات من نوع "{type_label}" لهذه الفئة.</strong><br>'
                f'<span style="font-size:11px">لكن يوجد <strong>{any_count}</strong> محاولة من أنواع أخرى.</span><br>'
                f'<br>📋 الأنواع المتاحة لهذه الفئة: <strong style="color:var(--blue)">{available_str}</strong><br>'
                f'<br><span style="color:var(--green)">💡 غيّري نوع الاختبار من القائمة أعلاه إلى أحد الأنواع المتاحة لمشاهدة التحليل.</span>'
            )

        return JsonResponse({
            'avg': None, 'attempts': 0, 'pass_pct': None, 'fail_pct': None,
            'highest': None, 'lowest': None, 'improvement': None,
            'descriptive': descriptive,
            'skills_mastered': 0, 'skills_needs': 0, 'skills': [], 'types': [],
            'no_data': True, 'available_types': available_types, 'any_count': any_count,
        })

    avg = round(qs.aggregate(a=Avg('percentage'))['a'] or 0, 1)
    pass_count = qs.filter(passed=True).count()
    pass_pct = round(100 * pass_count / total_qs)
    fail_pct = 100 - pass_pct
    high = round(max((float(r.percentage or 0) for r in qs), default=0), 1)
    low = round(min((float(r.percentage or 0) for r in qs), default=0), 1)

    # نسبة التحسن: متوسط النصف الأخير - متوسط النصف الأول
    sorted_qs = list(qs.order_by('submitted_at'))
    half = max(1, len(sorted_qs) // 2)
    first_half = sorted_qs[:half]
    second_half = sorted_qs[half:] if len(sorted_qs) > half else []
    avg_first = sum(float(r.percentage or 0) for r in first_half) / max(len(first_half), 1)
    avg_second = (sum(float(r.percentage or 0) for r in second_half) / max(len(second_half), 1)) if second_half else avg_first
    improvement = round(avg_second - avg_first, 1)

    skills, mastered, needs = _skills_breakdown(qs)
    types = _exam_types_breakdown(qs)

    descriptive = _descriptive_text(scope, avg, total_qs, pass_pct, improvement)

    return JsonResponse({
        'avg': avg, 'attempts': total_qs, 'pass_pct': pass_pct, 'fail_pct': fail_pct,
        'highest': high, 'lowest': low, 'improvement': improvement,
        'descriptive': descriptive,
        'skills_mastered': mastered, 'skills_needs': needs, 'skills': skills,
        'types': types,
    })


@admin_required
def admin_report(request):
    """يولّد تقريراً قابلاً للطباعة — طالبة/فصل/معلمة/مدرسة. آمن من الأخطاء."""
    try:
        return _admin_report_inner(request)
    except Exception as exc:
        # نلتقط أي خطأ غير متوقع ونعرضه بشكل ودود بدل صفحة 500
        import traceback
        return render(request, 'core/report.html', {
            'settings': SchoolSettings.objects.first(),
            'kind': request.GET.get('kind', ''),
            'audience': request.GET.get('audience', ''),
            'error': f'تعذّر إنشاء التقرير: {exc}',
            'trace_short': str(exc)[:200],
        })


def _admin_report_inner(request):
    kind = request.GET.get('kind', 'school')
    audience = request.GET.get('audience', 'admin')   # student / parent / admin
    target_id = (request.GET.get('id') or '').strip()

    settings_obj = SchoolSettings.objects.first()
    ctx = {
        'settings': settings_obj,
        'kind': kind,
        'audience': audience,
    }

    if kind == 'student':
        if not target_id or not target_id.isdigit():
            return render(request, 'core/report.html', {**ctx, 'error': 'يرجى اختيار الطالبة قبل إنشاء التقرير'})
        try:
            user = User.objects.get(id=int(target_id))
        except (User.DoesNotExist, ValueError):
            user = None
        if user is None:
            return render(request, 'core/report.html', {**ctx, 'error': 'الطالبة غير موجودة'})
        qs = ExamResult.objects.filter(student=user).select_related('exam', 'exam__skill')
        skills, mastered, needs = _skills_breakdown(qs)
        types = _exam_types_breakdown(qs)
        try:
            profile = user.core_profile
        except Profile.DoesNotExist:
            profile = None
        ctx.update({
            'student': user, 'profile': profile,
            'results': qs.order_by('-submitted_at'),
            'avg': round(qs.aggregate(a=Avg('percentage'))['a'] or 0, 1),
            'attempts': qs.count(),
            'passed': qs.filter(passed=True).count(),
            'failed': qs.filter(passed=False).count(),
            'skills': skills, 'mastered': mastered, 'needs': needs,
            'types': types,
            'descriptive': _descriptive_text('student',
                                             round(qs.aggregate(a=Avg('percentage'))['a'] or 0, 1),
                                             qs.count(),
                                             round(100 * qs.filter(passed=True).count() / max(qs.count(), 1)),
                                             0),
        })

    elif kind == 'classroom':
        cls_name = target_id
        if not cls_name:
            return render(request, 'core/report.html', {**ctx, 'error': 'يرجى اختيار الفصل'})
        names = list(Student.objects.filter(classroom__name=cls_name).values_list('full_name', flat=True))
        from django.db.models import Q
        cond = Q()
        valid = False
        for n in names:
            parts = n.split()
            if parts and parts[0]:
                cond |= Q(student__first_name=parts[0]) | Q(student__username=n)
                valid = True
        qs = ExamResult.objects.filter(cond) if valid else ExamResult.objects.none()
        skills, mastered, needs = _skills_breakdown(qs)
        types = _exam_types_breakdown(qs)
        ctx.update({
            'classroom_name': cls_name,
            'students_count': len(names),
            'avg': round(qs.aggregate(a=Avg('percentage'))['a'] or 0, 1),
            'attempts': qs.count(),
            'passed': qs.filter(passed=True).count(),
            'skills': skills, 'mastered': mastered, 'needs': needs, 'types': types,
            'descriptive': _descriptive_text('classroom',
                                             round(qs.aggregate(a=Avg('percentage'))['a'] or 0, 1),
                                             qs.count(),
                                             round(100 * qs.filter(passed=True).count() / max(qs.count(), 1)),
                                             0),
        })

    elif kind == 'teacher':
        if not target_id or not target_id.isdigit():
            return render(request, 'core/report.html', {**ctx, 'error': 'يرجى اختيار المعلمة'})
        # القائمة في الـ dashboard ترسل User.id (وليس Teacher.id) — نتعامل مع كليهما
        tid = int(target_id)
        teacher = (Teacher.objects.filter(user_id=tid).first()
                   or Teacher.objects.filter(id=tid).first())
        if teacher is None:
            return render(request, 'core/report.html', {**ctx, 'error': 'المعلمة غير موجودة'})
        qs = ExamResult.objects.filter(exam__skill__created_by=teacher)
        skills_count = TeacherSkill.objects.filter(created_by=teacher).count()
        sessions_count = ClassSession.objects.filter(teacher=teacher).count()
        exams_count = TeacherExam.objects.filter(skill__created_by=teacher).count()
        avg = round(qs.aggregate(a=Avg('percentage'))['a'] or 0, 1)
        if avg >= 70:
            tier = ('عالية التأثير', 'var(--green)')
        elif avg >= 50:
            tier = ('متوسطة التأثير', 'var(--orange)')
        else:
            tier = ('تحتاج دعم', 'var(--red)')
        ctx.update({
            'teacher': teacher,
            'skills_count': skills_count, 'sessions_count': sessions_count, 'exams_count': exams_count,
            'avg': avg, 'attempts': qs.count(),
            'students_engaged': qs.values('student').distinct().count(),
            'tier': tier[0], 'tier_color': tier[1],
            'top_results': qs.order_by('-percentage')[:10],
        })

    else:  # school
        all_results = ExamResult.objects
        ctx.update({
            'total_students': Profile.objects.filter(role='STUDENT').count(),
            'total_teachers': Profile.objects.filter(role='TEACHER').count(),
            'total_classrooms': ClassRoom.objects.count(),
            'total_skills': TeacherSkill.objects.count(),
            'total_exams': TeacherExam.objects.count(),
            'avg': round(all_results.aggregate(a=Avg('percentage'))['a'] or 0, 1),
            'attempts': all_results.count(),
            'passed': all_results.filter(passed=True).count(),
            'classrooms_perf': _v2_classroom_compare(),
            'top_students': _v2_top_students(limit=10),
            'teachers_impact': _v2_teachers_impact(),
            'levels': _v2_level_distribution(),
        })

    return render(request, 'core/report.html', ctx)


# ─── إدارة المستخدمين ─────────────────────────────────────────

@admin_required
@require_POST
def admin_add_teacher(request):
    full_name = (request.POST.get('full_name') or '').strip()
    # نحفظ رقم الهوية كما كتبتْه المديرة بالضبط (عربي أو لاتيني).
    # المطابقة لاحقاً تجرّب الصيغتين معاً.
    national_id = (request.POST.get('national_id') or '').strip()
    pin_code = (request.POST.get('pin_code') or '').strip()

    if not full_name or not national_id:
        messages.error(request, 'يرجى إدخال الاسم ورقم الهوية')
        return redirect('admin_dashboard')
    if not _is_valid_id(national_id):
        messages.error(request, 'رقم هوية غير صالح (8-12 رقم)')
        return redirect('admin_dashboard')
    if pin_code and not pin_code.translate(_AR_TO_LA).isdigit():
        messages.error(request, 'رمز PIN يجب أن يكون أرقاماً')
        return redirect('admin_dashboard')
    # نمنع التكرار سواء أُدخلت الهوية بأرقام عربية أم لاتينية
    variants = _id_variants(national_id)
    if (User.objects.filter(username__in=variants).exists()
            or Profile.objects.filter(national_id__in=variants).exists()):
        messages.error(request, 'رقم الهوية مسجّل مسبقاً')
        return redirect('admin_dashboard')

    name_parts = full_name.split()
    user = User.objects.create_user(
        username=national_id,  # نحفظ كما كُتب
        password=pin_code or national_id,
        first_name=name_parts[0] if name_parts else full_name,
        last_name=' '.join(name_parts[1:]) if len(name_parts) > 1 else '',
    )
    Profile.objects.create(
        user=user,
        role='TEACHER',
        national_id=national_id,  # نحفظ كما كُتب
        pin_code=pin_code,
    )
    Teacher.objects.create(user=user, full_name=full_name)
    messages.success(request, f'✅ تم إضافة المعلمة {full_name}')
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_add_student(request):
    full_name = (request.POST.get('full_name') or '').strip()
    # نحفظ رقم الهوية كما كتبتْه المديرة (الأرقام العربية تبقى عربية).
    national_id = (request.POST.get('national_id') or '').strip()
    classroom_id = (request.POST.get('classroom_id') or '').strip()

    if not full_name or not national_id:
        messages.error(request, 'يرجى إدخال الاسم ورقم الهوية')
        return redirect('admin_dashboard')
    if not _is_valid_id(national_id):
        messages.error(request, 'رقم هوية غير صالح (8-12 رقم)')
        return redirect('admin_dashboard')
    variants = _id_variants(national_id)
    if (User.objects.filter(username__in=variants).exists()
            or Profile.objects.filter(national_id__in=variants).exists()):
        messages.error(request, 'رقم الهوية مسجّل مسبقاً')
        return redirect('admin_dashboard')

    name_parts = full_name.split()
    user = User.objects.create_user(
        username=national_id,  # نحفظ كما كُتب (عربي/لاتيني)
        password=national_id,
        first_name=name_parts[0] if name_parts else full_name,
        last_name=' '.join(name_parts[1:]) if len(name_parts) > 1 else '',
    )
    # PIN فارغ للطالبة → دخول بالهوية فقط
    Profile.objects.create(
        user=user,
        role='STUDENT',
        national_id=national_id,
        pin_code='',
    )

    # classroom_id قد يكون فارغاً — نتعامل بأمان
    if classroom_id and classroom_id.isdigit():
        classroom = ClassRoom.objects.filter(id=classroom_id).first()
        if classroom:
            student_obj, _ = Student.objects.get_or_create(
                full_name=full_name,
                defaults={'classroom': classroom},
            )
            # ربط سجل الطالبة بحسابها تلقائياً
            if student_obj.user_id is None:
                student_obj.user = user
                student_obj.save(update_fields=['user'])
    messages.success(request, f'✅ تم إضافة الطالبة {full_name} (رقم الهوية: {national_id})')
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_delete_user(request, user_id):
    """حذف مستخدم — مع حماية الحساب الحالي و superuser."""
    target = User.objects.filter(id=user_id).first()
    if target is None:
        messages.error(request, 'المستخدم غير موجود')
        return redirect('admin_dashboard')
    if target.id == request.user.id:
        messages.error(request, 'لا يمكنك حذف حسابك أثناء الجلسة')
        return redirect('admin_dashboard')
    if target.is_superuser:
        messages.error(request, 'لا يمكن حذف حساب superuser')
        return redirect('admin_dashboard')

    name = target.get_full_name() or target.username

    # حذف سجل Student المرتبط إن كانت الطالبة (يجب قبل حذف الـ User)
    try:
        role = target.core_profile.role
        if role == 'STUDENT' and name:
            Student.objects.filter(full_name=name).delete()
    except Exception:
        pass

    target.delete()
    messages.success(request, f'🗑️ تم حذف {name}')
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_edit_user(request, user_id):
    """تعديل بيانات مستخدم: الاسم + PIN + تفعيل/تعطيل."""
    target = get_object_or_404(User, id=user_id)
    if target.is_superuser and not request.user.is_superuser:
        messages.error(request, 'لا يمكن تعديل حساب superuser')
        return redirect('admin_dashboard')

    full_name = (request.POST.get('full_name') or '').strip()
    pin_code = (request.POST.get('pin_code') or '').strip()
    action = request.POST.get('action') or 'save'

    if action == 'toggle':
        target.is_active = not target.is_active
        target.save(update_fields=['is_active'])
        messages.success(request, f"✅ تم {'تفعيل' if target.is_active else 'تعطيل'} الحساب")
        return redirect('admin_dashboard')

    if full_name:
        parts = full_name.split()
        target.first_name = parts[0] if parts else full_name
        target.last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''
        target.save(update_fields=['first_name', 'last_name'])
        # نحدّث Teacher.full_name إن وُجد
        try:
            target.teacher.full_name = full_name
            target.teacher.save(update_fields=['full_name'])
        except Teacher.DoesNotExist:
            pass

    try:
        profile = target.core_profile
        if pin_code:
            if not pin_code.isdigit():
                messages.error(request, 'رمز PIN يجب أن يكون أرقاماً')
                return redirect('admin_dashboard')
            profile.pin_code = pin_code
        # نسمح بمسح PIN عبر إرسال "clear"
        if request.POST.get('clear_pin') == '1':
            profile.pin_code = ''
        profile.save(update_fields=['pin_code'])
    except Profile.DoesNotExist:
        pass

    messages.success(request, '✅ تم حفظ التعديلات')
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_add_classroom(request):
    """إضافة فصل دراسي."""
    name = (request.POST.get('name') or '').strip()
    if not name:
        messages.error(request, 'يرجى إدخال اسم الفصل')
        return redirect('admin_dashboard')
    cls, created = ClassRoom.objects.get_or_create(name=name)
    if created:
        messages.success(request, f'✅ تم إضافة الفصل {name}')
    else:
        messages.warning(request, f'⚠️ الفصل {name} موجود مسبقاً')
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_delete_classroom(request, classroom_id):
    cls = get_object_or_404(ClassRoom, id=classroom_id)
    name = cls.name
    cls.delete()
    messages.success(request, f'🗑️ تم حذف الفصل {name}')
    return redirect('admin_dashboard')


# ─── استيراد Excel للمعلمات ───────────────────────────────────

MAX_EXCEL_SIZE = 5 * 1024 * 1024


def _read_excel_or_msg(request, redirect_target):
    """يقرأ ملف Excel من الطلب ويرجع worksheet أو يعيد توجيه برسالة خطأ."""
    excel = request.FILES.get('excel_file')
    if not excel:
        messages.error(request, 'لم يُرفع ملف')
        return None
    if excel.size > MAX_EXCEL_SIZE:
        messages.error(request, 'حجم الملف يتجاوز 5 ميجا')
        return None
    if not excel.name.lower().endswith(('.xlsx', '.xlsm')):
        messages.error(request, 'الصيغة المدعومة: xlsx/xlsm')
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel, data_only=True)
        return wb.active
    except Exception as exc:
        messages.error(request, f'❌ تعذّر فتح الملف: {exc}')
        return None


@admin_required
@require_POST
def admin_import_teachers(request):
    """
    استيراد معلمات من Excel.
    أعمدة الملف المتوقعة: الاسم | رقم الهوية | PIN (اختياري)
    """
    ws = _read_excel_or_msg(request, 'admin_dashboard')
    if ws is None:
        return redirect('admin_dashboard')

    count, skipped = 0, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        full_name = str(row[0]).strip() if row[0] else ''
        national_id = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        pin = str(row[2]).strip() if len(row) > 2 and row[2] else ''

        if not full_name or not national_id:
            continue
        if not _is_valid_id(national_id):
            skipped.append(f"{full_name}: هوية غير صالحة")
            continue
        variants = _id_variants(national_id)
        if (User.objects.filter(username__in=variants).exists()
                or Profile.objects.filter(national_id__in=variants).exists()):
            skipped.append(f"{full_name}: مكرّر")
            continue

        try:
            parts = full_name.split()
            user = User.objects.create_user(
                username=national_id,
                password=pin or national_id,
                first_name=parts[0] if parts else full_name,
                last_name=' '.join(parts[1:]) if len(parts) > 1 else '',
            )
            Profile.objects.create(
                user=user, role='TEACHER',
                national_id=national_id, pin_code=pin,
            )
            Teacher.objects.create(user=user, full_name=full_name)
            count += 1
        except Exception as exc:
            skipped.append(f"{full_name}: {exc}")

    if skipped:
        messages.warning(
            request,
            f'✅ استُورد {count} معلمة • تم تخطّي {len(skipped)}: ' + ' | '.join(skipped[:5])
        )
    else:
        messages.success(request, f'✅ تم استيراد {count} معلمة')
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_import_students(request):
    """
    استيراد طالبات من Excel.
    أعمدة الملف المتوقعة: الاسم | رقم الهوية | الفصل
    """
    ws = _read_excel_or_msg(request, 'admin_dashboard')
    if ws is None:
        return redirect('admin_dashboard')

    count, skipped = 0, []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        full_name = str(row[0]).strip() if row[0] else ''
        # تحويل رقم الهوية: Excel يقرأ الأرقام كـ float (مثل 1158153922.0)
        raw_id = row[1] if len(row) > 1 and row[1] is not None else ''
        if isinstance(raw_id, float):
            national_id = str(int(raw_id))
        else:
            national_id = str(raw_id).strip()
        classroom_name = str(row[2]).strip() if len(row) > 2 and row[2] else 'ث١٢'

        if not full_name or not national_id:
            continue
        if not _is_valid_id(national_id):
            skipped.append(f"{full_name}: هوية غير صالحة")
            continue
        variants = _id_variants(national_id)
        if (User.objects.filter(username__in=variants).exists()
                or Profile.objects.filter(national_id__in=variants).exists()):
            skipped.append(f"{full_name}: مكرّر")
            continue

        try:
            classroom, _ = ClassRoom.objects.get_or_create(name=classroom_name)
            parts = full_name.split()
            user = User.objects.create_user(
                username=national_id,
                password=national_id,
                first_name=parts[0] if parts else full_name,
                last_name=' '.join(parts[1:]) if len(parts) > 1 else '',
            )
            Profile.objects.create(
                user=user, role='STUDENT',
                national_id=national_id, pin_code='',
            )
            st_obj, _ = Student.objects.get_or_create(
                full_name=full_name,
                defaults={'classroom': classroom},
            )
            # ربط سجل الطالبة بحسابها تلقائياً
            if st_obj.user_id is None:
                st_obj.user = user
                st_obj.save(update_fields=['user'])
            count += 1
        except Exception as exc:
            skipped.append(f"{full_name}: {exc}")

    if skipped:
        messages.warning(
            request,
            f'✅ استُورد {count} طالبة • تم تخطّي {len(skipped)}: ' + ' | '.join(skipped[:5])
        )
    else:
        messages.success(request, f'✅ تم استيراد {count} طالبة')
    return redirect('admin_dashboard')


# ─── الدخول كمستخدم ──────────────────────────────────────────

@admin_required
@require_POST
def admin_view_as(request, user_id):
    target = get_object_or_404(User, id=user_id)
    target_role = _get_role(target)

    if target_role == 'ADMIN' or target.is_superuser:
        messages.error(request, 'لا يمكن الدخول كحساب إدارة آخر')
        return redirect('admin_dashboard')

    request.session['admin_id'] = request.user.id
    auth_login(request, target, backend='django.contrib.auth.backends.ModelBackend')

    if target_role == 'TEACHER':
        return redirect('teacher_dashboard')
    if target_role == 'STUDENT':
        return redirect('student_dashboard')
    return redirect('home')


@login_required
@require_POST
def admin_return(request):
    admin_id = request.session.get('admin_id')
    if not admin_id:
        return redirect('home')
    try:
        admin_user = User.objects.get(id=admin_id)
    except User.DoesNotExist:
        return redirect('home')

    auth_login(request, admin_user, backend='django.contrib.auth.backends.ModelBackend')
    request.session.pop('admin_id', None)
    return redirect('admin_dashboard')


# ─── الاختبارات الشاملة ──────────────────────────────────────

@admin_required
def admin_comprehensive(request):
    """قائمة الاختبارات الشاملة (قدرات/تحصيلي)."""
    comp_skills = (
        TeacherSkill.objects
        .filter(content_type='comprehensive')
        .select_related('created_by')
        .prefetch_related('exams')
        .order_by('-created_at')
    )
    return render(request, 'core/admin_comprehensive.html', {
        'comp_skills': comp_skills,
        'classrooms': ClassRoom.objects.all().order_by('name'),
    })


@admin_required
@require_POST
def admin_add_comprehensive(request):
    """إنشاء اختبار شامل جديد."""
    title = (request.POST.get('title') or '').strip()
    comp_type = (request.POST.get('comp_type') or '').strip()

    if not title:
        messages.error(request, 'يرجى إدخال عنوان الاختبار')
        return redirect('admin_comprehensive')
    if comp_type not in ('qodrat_kamy', 'tahsili'):
        messages.error(request, 'يرجى اختيار نوع الاختبار (قدرات/تحصيلي)')
        return redirect('admin_comprehensive')

    duration = _safe_int(request.POST.get('duration'), default=120, minimum=1, maximum=600)
    questions_count = _safe_int(request.POST.get('questions_count'), default=60, minimum=1, maximum=300)
    pass_score = _safe_int(request.POST.get('pass_score'), default=50, minimum=0, maximum=100)

    teacher, _ = Teacher.objects.get_or_create(
        user=request.user,
        defaults={'full_name': request.user.get_full_name() or 'المديرة'},
    )

    skill = TeacherSkill.objects.create(
        content_type='comprehensive',
        title=title,
        skill_type=comp_type,
        description=request.POST.get('description', ''),
        created_by=teacher,
        target_classes='جميع الفصول',
        is_active=False,
    )
    TeacherExam.objects.create(
        skill=skill,
        exam_type='comprehensive_qodrat' if comp_type == 'qodrat_kamy' else 'comprehensive_tahsili',
        questions_count=questions_count,
        duration_minutes=duration,
        pass_score=pass_score,
        is_active=False,
    )
    messages.success(request, f'✅ تم إنشاء "{title}" — يمكنك إضافة الأسئلة من زر "الأسئلة".')
    return redirect('admin_comprehensive')


@admin_required
@require_POST
def admin_edit_comprehensive(request, skill_id):
    """تعديل اختبار شامل (العنوان/المدة/الأسئلة/درجة النجاح)."""
    skill = get_object_or_404(TeacherSkill, pk=skill_id, content_type='comprehensive')
    exam = skill.exams.first()

    title = (request.POST.get('title') or '').strip()
    if title:
        skill.title = title
    skill.description = request.POST.get('description', skill.description)
    skill.save(update_fields=['title', 'description'])

    if exam:
        exam.questions_count = _safe_int(request.POST.get('questions_count'), exam.questions_count, minimum=1, maximum=300)
        exam.duration_minutes = _safe_int(request.POST.get('duration'), exam.duration_minutes, minimum=1, maximum=600)
        exam.pass_score = _safe_int(request.POST.get('pass_score'), exam.pass_score, minimum=0, maximum=100)
        exam.save(update_fields=['questions_count', 'duration_minutes', 'pass_score'])

    messages.success(request, '✅ تم حفظ التعديلات')
    return redirect('admin_comprehensive')


@admin_required
@require_POST
def admin_delete_comprehensive(request, skill_id):
    """حذف اختبار شامل (يحذف الأسئلة + النتائج المرتبطة بفعل CASCADE)."""
    skill = get_object_or_404(TeacherSkill, pk=skill_id, content_type='comprehensive')
    title = skill.title
    skill.delete()
    messages.success(request, f'🗑️ تم حذف "{title}"')
    return redirect('admin_comprehensive')


# ═══════════════════════════════════════════════════════════════
# إدارة أسئلة الاختبار الشامل (للمديرة)
#   - عرض جميع الأسئلة في صفحة واحدة
#   - إدخال يدوي مع لوحة رموز رياضية
#   - رفع Excel مباشرة من نفس الصفحة
#   - تعديل/حذف لكل سؤال
# ═══════════════════════════════════════════════════════════════

def _get_or_create_comp_exam(skill):
    """يضمن وجود exam مرتبط بالاختبار الشامل."""
    exam = skill.exams.first()
    if exam:
        return exam
    return TeacherExam.objects.create(
        skill=skill,
        exam_type='comprehensive_qodrat' if skill.skill_type == 'qodrat_kamy' else 'comprehensive_tahsili',
        questions_count=skill.exams.count() or 60,
        duration_minutes=120,
        pass_score=50,
        is_active=False,
    )


@admin_required
def admin_comp_questions(request, skill_id):
    """صفحة إدارة الأسئلة الكاملة لاختبار شامل."""
    skill = get_object_or_404(TeacherSkill, pk=skill_id, content_type='comprehensive')
    exam = _get_or_create_comp_exam(skill)
    questions = exam.questions.all().order_by('order')
    return render(request, 'core/admin_comp_questions.html', {
        'skill': skill,
        'exam': exam,
        'questions': questions,
    })


@admin_required
@require_POST
def admin_comp_add_question(request, skill_id):
    """إضافة سؤال يدوياً (مع دعم الرموز الرياضية في النص)."""
    skill = get_object_or_404(TeacherSkill, pk=skill_id, content_type='comprehensive')
    exam = _get_or_create_comp_exam(skill)

    text = (request.POST.get('question_plain') or '').strip()
    if not text:
        messages.error(request, 'يرجى إدخال نص السؤال')
        return redirect('admin_comp_questions', skill_id=skill.id)

    next_order = (exam.questions.count() or 0) + 1
    q = TeacherQuestion.objects.create(
        exam=exam,
        order=next_order,
        question_plain=text,
        option_a_plain=request.POST.get('option_a_plain', '').strip(),
        option_b_plain=request.POST.get('option_b_plain', '').strip(),
        option_c_plain=request.POST.get('option_c_plain', '').strip(),
        option_d_plain=request.POST.get('option_d_plain', '').strip(),
        correct_answer=(request.POST.get('correct_answer', 'A') or 'A').upper()[:1],
        target_skill_name=request.POST.get('target_skill_name', '').strip(),
        feedback_plain=request.POST.get('feedback_plain', '').strip(),
    )
    # صور: السؤال + الخيارات (يدعم data: URL من المسح الضوئي + ملفات مرفوعة)
    _attach_image(q, 'question_image', request)
    for letter in ('a', 'b', 'c', 'd'):
        _attach_image(q, f'option_{letter}_image', request)
    q.save()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'ok': True, 'order': next_order})
    messages.success(request, f'✅ تم إضافة السؤال رقم {next_order}')
    return redirect('admin_comp_questions', skill_id=skill.id)


def _attach_image(obj, field_name, request):
    """يرفع الصورة مباشرة على Cloudinary ويخزّن الـ public_id.
    يدعم: ملف مرفوع عادي أو data:URL من الرسم/OCR.
    يُرجع True عند النجاح، False عند الفشل، None إذا لا يوجد ملف."""
    import base64
    import cloudinary.uploader

    file_obj = None

    # 1) ملف مرفوع عادي
    if field_name in request.FILES:
        file_obj = request.FILES[field_name]
        logger.info(f"📎 Found file upload for {field_name}: {file_obj.name} ({file_obj.size} bytes)")

    # 2) data: URL (للصور الملصقة من الرسم/OCR)
    if not file_obj:
        data_url = request.POST.get(field_name + '_data') or request.POST.get(field_name + '_dataurl')
        if data_url and data_url.startswith('data:image'):
            file_obj = data_url  # Cloudinary يقبل data URLs مباشرة
            logger.info(f"📎 Found data URL for {field_name} (length={len(data_url)})")

    if not file_obj:
        return None

    try:
        folder = 'questions/options' if 'option_' in field_name else 'questions'
        result = cloudinary.uploader.upload(file_obj, folder=folder)
        public_id = result['public_id']
        setattr(obj, field_name, public_id)
        logger.info(f"✅ Cloudinary upload OK: {field_name} → {public_id} (url={result.get('secure_url','')})")
        return True
    except Exception as exc:
        logger.error(f"❌ Cloudinary upload FAILED for {field_name}: {exc}", exc_info=True)
        return False


@admin_required
@require_POST
def admin_comp_import_excel(request, skill_id):
    """استيراد أسئلة الاختبار الشامل من Excel."""
    skill = get_object_or_404(TeacherSkill, pk=skill_id, content_type='comprehensive')
    exam = _get_or_create_comp_exam(skill)

    excel = request.FILES.get('excel_file')
    if not excel:
        messages.error(request, 'لم يُرفع ملف')
        return redirect('admin_comp_questions', skill_id=skill.id)
    if excel.size > MAX_EXCEL_SIZE:
        messages.error(request, 'حجم الملف يتجاوز 5 ميجا')
        return redirect('admin_comp_questions', skill_id=skill.id)
    if not excel.name.lower().endswith(('.xlsx', '.xlsm')):
        messages.error(request, 'الصيغة المدعومة: xlsx/xlsm فقط')
        return redirect('admin_comp_questions', skill_id=skill.id)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel, data_only=True)
        ws = wb.active
    except Exception as exc:
        messages.error(request, f'❌ تعذّر فتح الملف: {exc}')
        return redirect('admin_comp_questions', skill_id=skill.id)

    base_order = exam.questions.count() or 0
    new_questions = []
    skipped = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row or not row[0]:
            continue
        q_text = str(row[0]).strip() if row[0] else ''
        if not q_text:
            skipped += 1
            continue
        new_questions.append(TeacherQuestion(
            exam=exam,
            order=base_order + i,
            question_plain=q_text,
            option_a_plain=str(row[1]).strip() if len(row) > 1 and row[1] else '',
            option_b_plain=str(row[2]).strip() if len(row) > 2 and row[2] else '',
            option_c_plain=str(row[3]).strip() if len(row) > 3 and row[3] else '',
            option_d_plain=str(row[4]).strip() if len(row) > 4 and row[4] else '',
            correct_answer=(str(row[5]).strip().upper()[:1] if len(row) > 5 and row[5] else 'A'),
            target_skill_name=str(row[6]).strip() if len(row) > 6 and row[6] else '',
            feedback_plain=str(row[7]).strip() if len(row) > 7 and row[7] else '',
        ))
    if new_questions:
        TeacherQuestion.objects.bulk_create(new_questions)
        messages.success(request, f'✅ تم استيراد {len(new_questions)} سؤال')
    else:
        messages.warning(request, '⚠️ لم يُستورد أي سؤال — تأكدي من تنسيق الملف')
    return redirect('admin_comp_questions', skill_id=skill.id)


@admin_required
@require_POST
def admin_comp_delete_question(request, skill_id, q_id):
    skill = get_object_or_404(TeacherSkill, pk=skill_id, content_type='comprehensive')
    exam = skill.exams.first()
    if exam:
        TeacherQuestion.objects.filter(id=q_id, exam=exam).delete()
        messages.success(request, '🗑️ تم حذف السؤال')
    return redirect('admin_comp_questions', skill_id=skill.id)


@admin_required
@require_POST
def admin_comp_edit_question(request, skill_id, q_id):
    skill = get_object_or_404(TeacherSkill, pk=skill_id, content_type='comprehensive')
    exam = skill.exams.first()
    if not exam:
        return redirect('admin_comp_questions', skill_id=skill.id)
    q = get_object_or_404(TeacherQuestion, pk=q_id, exam=exam)

    q.question_plain = (request.POST.get('question_plain') or q.question_plain).strip()
    q.option_a_plain = request.POST.get('option_a_plain', q.option_a_plain).strip()
    q.option_b_plain = request.POST.get('option_b_plain', q.option_b_plain).strip()
    q.option_c_plain = request.POST.get('option_c_plain', q.option_c_plain).strip()
    q.option_d_plain = request.POST.get('option_d_plain', q.option_d_plain).strip()
    q.correct_answer = (request.POST.get('correct_answer', q.correct_answer) or 'A').upper()[:1]
    q.target_skill_name = request.POST.get('target_skill_name', q.target_skill_name).strip()
    q.feedback_plain = request.POST.get('feedback_plain', q.feedback_plain).strip()
    # تحديث الصور (من ملف أو data URL)
    _attach_image(q, 'question_image', request)
    for letter in ('a', 'b', 'c', 'd'):
        _attach_image(q, f'option_{letter}_image', request)
    q.save()
    messages.success(request, f'✅ تم حفظ السؤال {q.order}')
    return redirect('admin_comp_questions', skill_id=skill.id)


# ═══════════════════════════════════════════════════════════════
# API: تحويل رسم المعادلة إلى LaTeX (Math OCR)
# ═══════════════════════════════════════════════════════════════

logger = logging.getLogger(__name__)


@login_required
@require_POST
def math_ocr(request):
    """
    يستقبل صورة Canvas (base64 PNG) ويحولها إلى LaTeX
    باستخدام Mathpix API.

    POST JSON: {"image": "<base64 png data>"}
    Response:  {"latex": "\\frac{3}{4}"} أو {"error": "..."}
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'invalid_json'}, status=400)

    image_b64 = body.get('image', '').strip()
    if not image_b64:
        return JsonResponse({'error': 'no_image'}, status=400)

    # ── التحقق من مفاتيح Mathpix ──
    app_id  = os.environ.get('MATHPIX_APP_ID', '').strip()
    app_key = os.environ.get('MATHPIX_APP_KEY', '').strip()

    if not app_id or not app_key:
        return JsonResponse({
            'error': 'not_configured',
            'latex': '',
            'message': 'Mathpix API keys not configured.'
        })

    # ── إرسال الطلب إلى Mathpix ──
    import urllib.request
    import urllib.error

    api_url = 'https://api.mathpix.com/v3/text'
    payload = json.dumps({
        'src': f'data:image/png;base64,{image_b64}',
        'formats': ['latex_simplified', 'asciimath'],
        'data_options': {
            'include_asciimath': True,
        }
    }).encode('utf-8')

    req = urllib.request.Request(api_url, data=payload, method='POST')
    req.add_header('Content-Type', 'application/json')
    req.add_header('app_id', app_id)
    req.add_header('app_key', app_key)

    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read().decode('utf-8'))

        latex = result.get('latex_simplified', '') or result.get('text', '')
        if latex:
            # تنظيف النتيجة
            latex = latex.strip()
            # إزالة أي $...$ خارجية إن وُجدت
            if latex.startswith('$') and latex.endswith('$'):
                latex = latex[1:-1].strip()
            if latex.startswith('\\(') and latex.endswith('\\)'):
                latex = latex[2:-2].strip()

            logger.info(f'Math OCR success: {latex[:60]}')
            return JsonResponse({'latex': latex})
        else:
            logger.warning(f'Math OCR empty result: {result}')
            return JsonResponse({
                'error': 'empty_result',
                'latex': '',
                'message': 'لم يتم التعرف على معادلة. حاولي الرسم بشكل أوضح.'
            })

    except urllib.error.HTTPError as e:
        body_err = e.read().decode('utf-8', errors='replace')
        logger.error(f'Mathpix HTTP error {e.code}: {body_err}')
        return JsonResponse({
            'error': 'api_error',
            'latex': '',
            'message': f'خطأ في Mathpix API ({e.code})'
        }, status=502)

    except Exception as e:
        logger.error(f'Math OCR exception: {e}')
        return JsonResponse({
            'error': 'server_error',
            'latex': '',
            'message': 'حدث خطأ في السيرفر.'
        }, status=500)

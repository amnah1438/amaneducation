from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db import models
from django.db.models import Avg, Count, Q
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from core.models import Profile
from students.models import Student, ClassRoom
from .models import (
    Teacher, TeacherSkill, TeacherSkillContent,
    TeacherExam, TeacherQuestion, ExamResult,
    StudentAnswer, ClassSession, SkillStandard
)


def get_teacher(request):
    """يرجع كائن Teacher المرتبط بالمستخدم، وإن لم يوجد ينشئه تلقائياً.
    سابقاً كان يرجع None إن لم يوجد، فيفشل add_skill_complete صامتاً."""
    try:
        return Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        # auto-heal: إن دخلت معلمة (لها Profile.role==TEACHER) لكن بلا Teacher row،
        # ننشئه فوراً بدل التسبب في 500 على الحفظ.
        try:
            profile = request.user.core_profile
            if profile.role == 'TEACHER':
                full_name = (
                    f"{request.user.first_name} {request.user.last_name}".strip()
                    or request.user.username
                )
                return Teacher.objects.create(user=request.user, full_name=full_name)
        except Profile.DoesNotExist:
            pass
        return None


def check_teacher(request):
    """يفحص أن المستخدم معلمة. لا نخفي الاستثناءات الفعلية."""
    try:
        return request.user.core_profile.role == 'TEACHER'
    except Profile.DoesNotExist:
        return False


def check_teacher_or_admin(request):
    try:
        role = request.user.core_profile.role
        return role in ('TEACHER', 'ADMIN')
    except Profile.DoesNotExist:
        return False


def _safe_int(value, default, minimum=None, maximum=None):
    """تحويل آمن للأرقام يمنع crash من حقول فارغة أو نص."""
    try:
        v = int(str(value).strip())
    except (TypeError, ValueError):
        return default
    if minimum is not None and v < minimum:
        return minimum
    if maximum is not None and v > maximum:
        return maximum
    return v


@never_cache
@login_required
def teacher_dashboard(request):
    """
    لوحة المعلمة الرئيسية — جميع الإحصاءات من قاعدة البيانات.

    الإصلاحات:
    - حساب excellent/mid/weak_students من ExamResult بدل البيانات المحفورة في القالب.
    - تجميع متوسط الأداء لكل طالبة (لا نتيجة فردية)، لتعطي صورة دقيقة.
    - إعادة sessions الأخيرة + المهارات الأخيرة بالقيم الحقيقية.
    """
    if not check_teacher(request):
        return redirect('home')

    teacher = get_teacher(request)
    if not teacher:
        # auto-heal: إذا لم يوجد Teacher record نُنشئه (انظر get_teacher)
        messages.warning(request, 'تم إنشاء حساب معلمة جديد لربطه ببياناتك.')
        return redirect('teacher_dashboard')

    # ─── فلاتر ───────────────────────────────────────────────
    f_classroom = request.GET.get('classroom', '')
    f_skill_type = request.GET.get('skill_type', '')

    # فصول المعلمة المفلترة — تُستخدم في كل مكان
    _teacher_classrooms = teacher.classrooms.all() if teacher else ClassRoom.objects.none()
    if f_classroom and f_classroom != 'all':
        _filtered_classrooms = _teacher_classrooms.filter(name=f_classroom)
    else:
        _filtered_classrooms = _teacher_classrooms

    # طالبات الفصل المفلتر (للتحقق من الغياب لاحقاً)
    _cls_qs = Student.objects.filter(classroom__in=_filtered_classrooms)

    my_skills = TeacherSkill.objects.filter(created_by__user=request.user)
    my_exams = TeacherExam.objects.filter(skill__created_by__user=request.user)
    my_results = ExamResult.objects.filter(exam__skill__created_by__user=request.user).exclude(student=request.user)

    # تطبيق فلتر نوع المهارة
    if f_skill_type and f_skill_type != 'all':
        type_map = {'qodrat': 'skill', 'tahsili': 'lesson'}
        mapped = type_map.get(f_skill_type, f_skill_type)
        my_skills = my_skills.filter(content_type=mapped)
        my_exams = my_exams.filter(skill__content_type=mapped)
        my_results = my_results.filter(exam__skill__content_type=mapped)

    # تطبيق فلتر الفصل — عبر Student→User (دقيق) + student_record
    if f_classroom and f_classroom != 'all':
        _user_ids    = _cls_qs.filter(user__isnull=False).values_list('user_id', flat=True)
        _record_ids  = _cls_qs.values_list('id', flat=True)
        my_results = my_results.filter(
            Q(student_id__in=_user_ids) | Q(student_record_id__in=_record_ids)
        )

    # ─── تجميع نتائج الطالبات (متوسط أداء كل طالبة) ───────────
    # نجمع النتائج من مصدرين: طالبات بحساب User + طالبات بسجل Student (رصد يدوي)
    from collections import defaultdict as _perf_dd
    _perf_map = _perf_dd(lambda: {'name': '', 'total_pct': 0, 'count': 0, 'key': ''})

    for r in my_results.select_related('student', 'student_record'):
        if r.student_record_id:
            key = f'sr_{r.student_record_id}'
            name = r.student_record.full_name
        elif r.student_id:
            key = f'u_{r.student_id}'
            u = r.student
            name = (u.get_full_name() or u.username) if u else '—'
        else:
            continue
        _perf_map[key]['name'] = name
        _perf_map[key]['total_pct'] += float(r.percentage or 0)
        _perf_map[key]['count'] += 1
        _perf_map[key]['key'] = key

    students_perf = []
    for k, v in _perf_map.items():
        if v['count'] > 0:
            avg = v['total_pct'] / v['count']
            students_perf.append({
                'name': v['name'], 'avg_pct': avg,
                'results_count': v['count'], 'key': k,
            })
    students_perf.sort(key=lambda x: -(x['avg_pct'] or 0))

    excellent_students = [
        {'name': s['name'], 'pct': int(round(s['avg_pct'] or 0)),
         'count': s['results_count'], 'key': s['key']}
        for s in students_perf if (s['avg_pct'] or 0) >= 90
    ]
    mid_students = [
        {'name': s['name'], 'pct': int(round(s['avg_pct'] or 0)),
         'count': s['results_count'], 'key': s['key']}
        for s in students_perf if 50 <= (s['avg_pct'] or 0) < 90
    ]
    weak_students = [
        {'name': s['name'], 'pct': int(round(s['avg_pct'] or 0)),
         'count': s['results_count'], 'key': s['key']}
        for s in students_perf if (s['avg_pct'] or 0) < 50
    ]

    avg_score = int(round(my_results.aggregate(avg=Avg('percentage'))['avg'] or 0))

    # ─── ترتيب أعلى 3 طالبات ──────────────────────────────────
    top_students = (excellent_students + mid_students)[:3]

    # ═══════════════════════════════════════════════════════════
    # بيانات حقيقية إضافية للويدجتس
    # ═══════════════════════════════════════════════════════════
    from datetime import timedelta as _td
    from collections import defaultdict as _dd
    import json as _json

    # 1) Heatmap — نشاط آخر 28 يوم (عدد المحاولات لكل يوم)
    today = timezone.now().date()
    heatmap_data = []
    for i in range(27, -1, -1):
        d = today - _td(days=i)
        cnt = my_results.filter(submitted_at__date=d).count()
        # مستوى 0-4 حسب الكثافة
        level = 0
        if cnt >= 10: level = 4
        elif cnt >= 5: level = 3
        elif cnt >= 2: level = 2
        elif cnt >= 1: level = 1
        heatmap_data.append({'date': d.isoformat(), 'count': cnt, 'level': level})

    # 2) المهارات الأكثر صعوبة (مفلترة بالفصل المختار)
    skill_stats = _dd(lambda: {'correct': 0, 'total': 0})
    _ans_qs = StudentAnswer.objects.filter(result__exam__skill__created_by=teacher).select_related('question')
    if f_classroom and f_classroom != 'all':
        _ans_qs = _ans_qs.filter(
            Q(result__student_id__in=_cls_qs.filter(user__isnull=False).values_list('user_id', flat=True)) |
            Q(result__student_record_id__in=_cls_qs.values_list('id', flat=True))
        )
    for ans in _ans_qs:
        name = (ans.question.target_skill_name or 'بدون تصنيف').strip() or 'بدون تصنيف'
        skill_stats[name]['total'] += 1
        if ans.is_correct: skill_stats[name]['correct'] += 1
    skill_hardness = []
    for name, v in skill_stats.items():
        if v['total'] >= 3:  # على الأقل 3 محاولات للموثوقية
            pct = round(100 * v['correct'] / v['total'])
            skill_hardness.append({'name': name, 'pct': pct, 'count': v['total']})
    skill_hardness.sort(key=lambda s: s['pct'])  # الأصعب أولاً
    skill_hardness = skill_hardness[:6]

    # 3) Radar — متوسط أداء أعلى 5 طالبات في كل قسم (sklill_type)
    radar_buckets = _dd(lambda: _dd(list))
    for r in my_results.select_related('exam', 'exam__skill', 'student'):
        section = r.exam.skill.skill_type or r.exam.skill.subject or 'عام'
        radar_buckets[r.student_id][section].append(float(r.percentage or 0))
    radar_data = []
    sections = sorted({s for d in radar_buckets.values() for s in d.keys()})
    for sid, secs in list(radar_buckets.items())[:5]:
        u = User.objects.filter(id=sid).first()
        name = (u.get_full_name() or u.username) if u else f'ID-{sid}'
        radar_data.append({
            'name': name,
            'values': [int(round(sum(secs.get(s, [0])) / max(len(secs.get(s, [])), 1))) for s in sections],
        })

    # 4) الطالبات الـ "مُعالَجَات" — تحسّن من قبلي إلى بعدي ≥ 10%
    pre_scores = {}
    post_scores = {}
    for r in my_results.select_related('exam', 'exam__skill'):
        sk_id = r.exam.skill_id
        key = (r.student_id, sk_id)
        if r.exam.exam_type == 'pre':
            pre_scores[key] = float(r.percentage or 0)
        elif r.exam.exam_type == 'post':
            post_scores[key] = float(r.percentage or 0)
    treated_count = 0
    for key, post in post_scores.items():
        pre = pre_scores.get(key)
        if pre is not None and (post - pre) >= 10:
            treated_count += 1

    # 5) عدد الطالبات الذين دربتهن (distinct students)
    trained_students = my_results.values('student_id').distinct().count()

    # 6) المهارات الأكثر طلباً للتدريب (مفلترة بالمعلمة والفصل)
    bank_gaps = _dd(int)
    _bank_qs = StudentAnswer.objects.filter(
        result__exam__exam_type='bank',
        result__exam__skill__created_by=teacher,
    ).select_related('question')
    if f_classroom and f_classroom != 'all':
        _bank_qs = _bank_qs.filter(
            Q(result__student_id__in=_cls_qs.filter(user__isnull=False).values_list('user_id', flat=True)) |
            Q(result__student_record_id__in=_cls_qs.values_list('id', flat=True))
        )
    for ans in _bank_qs:
        if not ans.is_correct and ans.question.target_skill_name:
            bank_gaps[ans.question.target_skill_name.strip()] += 1
    top_demand = sorted([(k, v) for k, v in bank_gaps.items()], key=lambda x: -x[1])[:5]

    # 7) skills_count breakdown
    skills_count = my_skills.filter(content_type='skill').count()
    lessons_count = my_skills.filter(content_type='lesson').count()
    banks_count = my_skills.filter(content_type='bank').count()

    # 8) بيانات مقارنة القبلي والبعدي لكل مهارة (للرسم البياني)
    pre_post_by_skill = _dd(lambda: {'pre_total': 0, 'pre_count': 0, 'post_total': 0, 'post_count': 0})
    for r in my_results.select_related('exam', 'exam__skill'):
        sk_name = r.exam.skill.title if r.exam.skill else 'بدون'
        if r.exam.exam_type == 'pre':
            pre_post_by_skill[sk_name]['pre_total'] += float(r.percentage or 0)
            pre_post_by_skill[sk_name]['pre_count'] += 1
        elif r.exam.exam_type == 'post':
            pre_post_by_skill[sk_name]['post_total'] += float(r.percentage or 0)
            pre_post_by_skill[sk_name]['post_count'] += 1
    pre_post_chart = []
    for sk_name, v in pre_post_by_skill.items():
        pre_avg = round(v['pre_total'] / v['pre_count']) if v['pre_count'] else 0
        post_avg = round(v['post_total'] / v['post_count']) if v['post_count'] else 0
        if v['pre_count'] or v['post_count']:
            pre_post_chart.append({'name': sk_name, 'pre': pre_avg, 'post': post_avg})

    # 9) بيانات الهيستوغرام — توزيع درجات الطالبات
    hist_bins = [0]*8  # 20-29, 30-39, 40-49, 50-59, 60-69, 70-79, 80-89, 90-100
    for sp in students_perf:
        pct = sp['avg_pct']
        if pct >= 90: hist_bins[7] += 1
        elif pct >= 80: hist_bins[6] += 1
        elif pct >= 70: hist_bins[5] += 1
        elif pct >= 60: hist_bins[4] += 1
        elif pct >= 50: hist_bins[3] += 1
        elif pct >= 40: hist_bins[2] += 1
        elif pct >= 30: hist_bins[1] += 1
        else: hist_bins[0] += 1

    # 10) ملخص المهارات — القبلي والبعدي والغيابات (accordion)
    # قائمة طالبات الفصل المفلتر مع id السجل + id الحساب (للتحقق من المصدرين)
    _cls_students = list(
        _cls_qs.values_list('id', 'full_name', 'user_id').order_by('full_name')
    ) if teacher else []

    def _took_set(exam):
        """يرجع مجموعة student IDs اللي أدّين هذا الاختبار — من أي مصدر."""
        if not exam:
            return set(), set()
        res = ExamResult.objects.filter(exam=exam)
        via_record = set(res.filter(student_record__isnull=False).values_list('student_record_id', flat=True))
        via_user   = set(res.filter(student__isnull=False).values_list('student_id', flat=True))
        return via_record, via_user

    def _missing(students, took_records, took_users, has_exam):
        if not has_exam:
            return []
        return [
            {'id': sid, 'name': sname}
            for sid, sname, suid in students
            if sid not in took_records and (suid is None or suid not in took_users)
        ]

    skills_summary = []
    _total_missing = 0
    _missing_skills_count = 0

    for _sk in my_skills.prefetch_related('exams'):
        _pre_exam  = _sk.exams.filter(exam_type='pre').first()
        _post_exam = _sk.exams.filter(exam_type='post').first()
        if not _pre_exam and not _post_exam:
            continue

        # القبلي
        _pre_rec, _pre_usr = _took_set(_pre_exam)
        _pre_avg = 0
        if _pre_exam:
            _pre_res = ExamResult.objects.filter(exam=_pre_exam)
            if f_classroom and f_classroom != 'all':
                _cls_uids = _cls_qs.filter(user__isnull=False).values_list('user_id', flat=True)
                _cls_rids = _cls_qs.values_list('id', flat=True)
                _pre_res = _pre_res.filter(
                    Q(student_id__in=_cls_uids) | Q(student_record_id__in=_cls_rids)
                )
            _pre_avg = int(round(_pre_res.aggregate(avg=Avg('percentage'))['avg'] or 0))

        # البعدي
        _post_rec, _post_usr = _took_set(_post_exam)
        _post_avg = 0
        if _post_exam:
            _post_res = ExamResult.objects.filter(exam=_post_exam)
            if f_classroom and f_classroom != 'all':
                _cls_uids = _cls_qs.filter(user__isnull=False).values_list('user_id', flat=True)
                _cls_rids = _cls_qs.values_list('id', flat=True)
                _post_res = _post_res.filter(
                    Q(student_id__in=_cls_uids) | Q(student_record_id__in=_cls_rids)
                )
            _post_avg = int(round(_post_res.aggregate(avg=Avg('percentage'))['avg'] or 0))

        _missing_pre  = _missing(_cls_students, _pre_rec,  _pre_usr,  bool(_pre_exam))
        _missing_post = _missing(_cls_students, _post_rec, _post_usr, bool(_post_exam))
        _pre_took_count  = len(_cls_students) - len(_missing_pre)  if _pre_exam  else 0
        _post_took_count = len(_cls_students) - len(_missing_post) if _post_exam else 0
        _improvement = (_post_avg - _pre_avg) if (_pre_exam and _post_exam and _pre_took_count) else None

        skills_summary.append({
            'skill_title': _sk.title,
            'skill_id': _sk.id,
            'has_pre': bool(_pre_exam),
            'has_post': bool(_post_exam),
            'pre_took': _pre_took_count,
            'pre_missing': len(_missing_pre),
            'pre_avg': _pre_avg,
            'post_took': _post_took_count,
            'post_missing': len(_missing_post),
            'post_avg': _post_avg,
            'improvement': _improvement,
            'missing_pre_students': _missing_pre,
            'missing_post_students': _missing_post,
        })
        if _missing_pre:
            _total_missing += len(_missing_pre)
            _missing_skills_count += 1

    missing_by_skill = [s for s in skills_summary if s['pre_missing'] > 0]

    return render(request, 'teachers/dashboard.html', {
        'teacher': teacher,
        'total_skills': skills_count,
        'total_lessons': lessons_count,
        'total_banks': banks_count,
        'active_skills': my_skills.filter(is_active=True).count(),
        'total_exams': my_exams.count(),
        'active_exams': my_exams.filter(is_active=True, results__in=my_results).distinct().count(),
        'avg_score': avg_score,
        'total_results': my_results.count(),
        'passed_results': my_results.filter(passed=True).count(),
        'total_students': _cls_qs.count() if teacher else 0,
        'trained_students': trained_students,        # طالبات دربتهن
        'treated_students': treated_count,           # طالبات عالجتهن (تحسّن ≥10%)
        'sessions': (
            ClassSession.objects
            .filter(teacher=teacher)
            .select_related('skill')
            .order_by('-session_date', '-session_time')[:10]
        ),
        'qodrat_sessions': ClassSession.objects.filter(
            teacher=teacher, session_type='qodrat',
            **({'target_class': f_classroom} if f_classroom and f_classroom != 'all' else {})
        ).count(),
        'tahsili_sessions': ClassSession.objects.filter(
            teacher=teacher, session_type='tahsili',
            **({'target_class': f_classroom} if f_classroom and f_classroom != 'all' else {})
        ).count(),
        'recent_skills': my_skills.order_by('-created_at')[:5],
        'excellent_students': excellent_students,
        'mid_students': mid_students,
        'weak_students': weak_students,
        'top_students': top_students,
        'students_count_excellent': len(excellent_students),
        'students_count_mid': len(mid_students),
        'students_count_weak': len(weak_students),
        'has_real_data': bool(students_perf),
        # ─── البيانات الحقيقية الجديدة ───
        'heatmap_json': _json.dumps(heatmap_data),
        'skill_hardness': skill_hardness,
        'radar_sections': sections,
        'radar_data_json': _json.dumps(radar_data),
        'top_demand_skills': [{'name': k, 'count': v} for k, v in top_demand],
        'pre_post_json': _json.dumps(pre_post_chart, ensure_ascii=False),
        'hist_bins_json': _json.dumps(hist_bins),
        'my_classrooms': teacher.classrooms.all().prefetch_related('students'),
        'my_exams_list': my_exams.select_related('skill'),
        # ─── بيانات الفلاتر ───
        'all_classrooms': teacher.classrooms.all().order_by('name'),
        'f_classroom': f_classroom,
        'f_skill_type': f_skill_type,
        # ─── ملخص المهارات accordion ───
        'skills_summary': skills_summary,
        'missing_by_skill': missing_by_skill,
        'total_missing_count': _total_missing,
        'missing_skills_count': _missing_skills_count,
    })


@login_required
def skill_manager(request):
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    my_skills = TeacherSkill.objects.filter(created_by=teacher).order_by('-created_at') if teacher else []
    shared_skills = TeacherSkill.objects.filter(is_shared=True, is_active=True).exclude(created_by=teacher).order_by('-created_at')[:20] if teacher else []
    # الفصول المعينة لهذه المعلمة فقط
    school_classrooms = teacher.classrooms.all().order_by('name') if teacher else ClassRoom.objects.none()
    context = {
        'teacher': teacher,
        'my_skills': my_skills,
        'shared_skills': shared_skills,
        'my_skills_count': my_skills.count() if teacher else 0,
        'school_classrooms': school_classrooms,
    }
    return render(request, 'teachers/skill_manager.html', context)


@login_required
def add_skill(request):
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    if not teacher:
        messages.error(request, 'لا يوجد حساب معلمة')
        return redirect('teacher_dashboard')
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        if not title:
            messages.error(request, 'يرجى إدخال عنوان المهارة')
            return redirect('skill_manager')
        skill = TeacherSkill.objects.create(
            content_type=request.POST.get('content_type', 'skill'),
            title=title,
            skill_type=request.POST.get('skill_type', ''),
            subject=request.POST.get('subject', ''),
            description=request.POST.get('description', ''),
            created_by=teacher,
            target_classes=request.POST.get('target_classes', ''),
            is_shared=request.POST.get('is_shared') == 'on',
            is_active=request.POST.get('is_active') == 'on',
        )
        video_url = request.POST.get('video_url', '')
        plain_text = request.POST.get('plain_text', '')
        if video_url or plain_text:
            TeacherSkillContent.objects.create(skill=skill, video_url=video_url, plain_text=plain_text)
        messages.success(request, f'تم إضافة "{title}" بنجاح!')
        return redirect('skill_manager')
    return render(request, 'teachers/skill_manager.html', {'teacher': teacher})


@login_required
def add_skill_complete(request):
    """
    حفظ مهارة/درس/بنك أسئلة كاملة دفعة واحدة (wizard 5 خطوات).

    إصلاحات حرجة:
    - _safe_int بدل int() لمنع crash عند إرسال حقل فارغ.
    - bulk_create للأسئلة لتقليل INSERTs من N إلى 1.
    - رسائل خطأ واضحة بدلاً من except صامت يخفي الأخطاء.
    - log إلى console بنوع المشكلة لتسهيل التشخيص.
    """
    import json
    import logging
    logger = logging.getLogger(__name__)

    if not check_teacher(request):
        return redirect('home')

    # === تشخيص: سجّل كل شيء يوصل من الفورم ===
    logger.info("=" * 60)
    logger.info(f"📥 add_skill_complete called")
    logger.info(f"📥 request.FILES keys: {list(request.FILES.keys())}")
    for k, v in request.FILES.items():
        logger.info(f"📥 FILE '{k}': name={v.name}, size={v.size}, content_type={v.content_type}")
    logger.info(f"📥 request.POST keys: {list(request.POST.keys())}")
    for k in ('video_url', 'plain_text', 'content_type', 'title'):
        logger.info(f"📥 POST '{k}': {repr(request.POST.get(k, '')[:100])}")
    logger.info("=" * 60)

    teacher = get_teacher(request)
    if not teacher:
        messages.error(request, 'لم يُربط حسابك بسجل معلمة. تواصلي مع المديرة.')
        return redirect('teacher_dashboard')

    if request.method != 'POST':
        return redirect('skill_manager')

    content_type = request.POST.get('content_type', 'skill')
    title = (request.POST.get('title') or '').strip()

    if not title:
        messages.error(request, 'يرجى إدخال عنوان المهارة')
        return redirect('skill_manager')

    is_active = request.POST.get('is_active') == 'on'

    # ─── إنشاء المهارة الأم ────────────────────────────────────
    try:
        skill = TeacherSkill.objects.create(
            content_type=content_type,
            title=title,
            skill_type=request.POST.get('skill_type', ''),
            subject=request.POST.get('subject', ''),
            description=request.POST.get('description', ''),
            created_by=teacher,
            target_classes=request.POST.get('target_classes', ''),
            is_shared=request.POST.get('is_shared') == 'on',
            is_active=is_active,
        )
    except Exception as e:
        logger.exception('Failed to create TeacherSkill')
        messages.error(request, f'تعذّر حفظ المهارة: {e}')
        return redirect('skill_manager')

    # ─── محتوى الشرح (اختياري) ────────────────────────────────
    import cloudinary.uploader
    video_url = request.POST.get('video_url', '').strip()
    plain_text = request.POST.get('plain_text', '').strip()
    plain_image = request.FILES.get('plain_image')
    pdf_file = request.FILES.get('pdf_file')
    if video_url or plain_text or plain_image or pdf_file:
        content = TeacherSkillContent(
            skill=skill, video_url=video_url, plain_text=plain_text,
        )
        if plain_image:
            try:
                logger.info(f"📎 Uploading skill image: {plain_image.name} ({plain_image.size} bytes)")
                r = cloudinary.uploader.upload(plain_image, folder='skill_content')
                content.plain_image = r['public_id']
                logger.info(f"✅ Skill image uploaded: {r['public_id']} → {r.get('secure_url','')}")
            except Exception as e:
                logger.error(f"❌ Image upload failed: {e}", exc_info=True)
                messages.warning(request, f'⚠️ فشل رفع الصورة: {e}')
        if pdf_file:
            max_pdf_mb = 10
            pdf_size_mb = round(pdf_file.size / (1024*1024), 1)
            if pdf_file.size > max_pdf_mb * 1024 * 1024:
                messages.warning(request, f'⚠️ ملف PDF كبير جداً ({pdf_size_mb} ميجا). الحد الأقصى {max_pdf_mb} ميجا. يرجى ضغط الملف وإعادة المحاولة.')
                logger.warning(f"PDF too large: {pdf_file.name} = {pdf_size_mb}MB (max {max_pdf_mb}MB)")
            else:
                try:
                    logger.info(f"📎 Uploading skill PDF: {pdf_file.name} ({pdf_size_mb}MB)")
                    r = cloudinary.uploader.upload(pdf_file, folder='skill_pdfs', resource_type='raw', access_mode='public')
                    content.pdf_file = r['public_id']
                    logger.info(f"✅ Skill PDF uploaded: {r['public_id']} → {r.get('secure_url','')}")
                except Exception as e:
                    logger.error(f"❌ PDF upload failed: {e}", exc_info=True)
                    messages.warning(request, f'⚠️ فشل رفع ملف PDF: {e}')
        try:
            content.save()
            logger.info(f"💾 Content saved OK: id={content.id}, plain_image={content.plain_image}, pdf_file={content.pdf_file}, video_url={content.video_url}")
            messages.info(request, f'📖 تم حفظ المحتوى (صورة: {"✅" if content.plain_image else "❌"}, PDF: {"✅" if content.pdf_file else "❌"}, فيديو: {"✅" if content.video_url else "❌"})')
        except Exception as e:
            logger.error(f"💥 Content save FAILED: {e}", exc_info=True)
            messages.error(request, f'⚠️ فشل حفظ المحتوى: {e}')

    # ─── الاختبارات والأسئلة ──────────────────────────────────
    def _bulk_questions(exam, raw_json, key_prefix=''):
        """يحفظ قائمة أسئلة من JSON بدون فشل صامت.
        يدعم الآن صور الأسئلة والخيارات كـ data:URL."""
        try:
            items = json.loads(raw_json or '[]')
        except json.JSONDecodeError as exc:
            logger.warning(f'Invalid {key_prefix} questions JSON: {exc}')
            return 0
        if not isinstance(items, list):
            return 0
        objs = []
        img_data = []  # قائمة بالصور المطلوب رفعها بعد الإنشاء
        for i, q in enumerate(items, start=1):
            if not isinstance(q, dict):
                continue
            # ربط المهارة المعيارية إن وُجدت
            ss_obj = None
            ss_id = q.get('skill_standard_id', '')
            if ss_id:
                try:
                    ss_obj = SkillStandard.objects.get(id=int(ss_id))
                except (SkillStandard.DoesNotExist, ValueError):
                    pass
            objs.append(TeacherQuestion(
                exam=exam,
                order=i,
                question_plain=q.get('text', ''),
                option_a_plain=q.get('a', ''),
                option_b_plain=q.get('b', ''),
                option_c_plain=q.get('c', ''),
                option_d_plain=q.get('d', ''),
                correct_answer=(q.get('correct') or 'A').upper()[:1],
                target_skill_name=q.get('skill', ''),
                skill_standard=ss_obj,
                feedback_plain=q.get('feedback', ''),
            ))
            # تجميع بيانات الصور (data:URL) إن وُجدت
            imgs = {}
            if q.get('q_img', '').startswith('data:image'):
                imgs['question_image'] = q['q_img']
            if q.get('a_img', '').startswith('data:image'):
                imgs['option_a_image'] = q['a_img']
            if q.get('b_img', '').startswith('data:image'):
                imgs['option_b_image'] = q['b_img']
            if q.get('c_img', '').startswith('data:image'):
                imgs['option_c_image'] = q['c_img']
            if q.get('d_img', '').startswith('data:image'):
                imgs['option_d_image'] = q['d_img']
            img_data.append(imgs)

        if objs:
            TeacherQuestion.objects.bulk_create(objs)

        # رفع الصور إلى Cloudinary بعد الإنشاء
        import cloudinary.uploader
        for idx, imgs in enumerate(img_data):
            if not imgs:
                continue
            q_obj = objs[idx]
            # بعد bulk_create يكون للعنصر pk
            if not q_obj.pk:
                # في بعض قواعد البيانات bulk_create لا يرجع pk
                try:
                    q_obj = TeacherQuestion.objects.get(exam=exam, order=idx+1)
                except TeacherQuestion.DoesNotExist:
                    continue
            for field_name, data_url in imgs.items():
                try:
                    folder = 'questions/options' if 'option_' in field_name else 'questions'
                    result = cloudinary.uploader.upload(data_url, folder=folder)
                    setattr(q_obj, field_name, result['public_id'])
                    logger.info(f"✅ Wizard image upload: {field_name} → {result['public_id']}")
                except Exception as e:
                    logger.error(f"❌ Wizard image upload FAILED for {field_name}: {e}")
            q_obj.save()

        return len(objs)

    pre_count = _safe_int(request.POST.get('pre_count'), default=10, minimum=1, maximum=100)
    pre_time = _safe_int(request.POST.get('pre_time'), default=15, minimum=1, maximum=300)
    pre_pass = _safe_int(request.POST.get('pre_pass'), default=60, minimum=0, maximum=100)
    post_count = _safe_int(request.POST.get('post_count'), default=10, minimum=1, maximum=100)
    post_time = _safe_int(request.POST.get('post_time'), default=15, minimum=1, maximum=300)
    post_pass = _safe_int(request.POST.get('post_pass'), default=70, minimum=0, maximum=100)

    qs_saved = 0
    if content_type == 'skill':
        # مهارة قدرات → اختباران (قبلي + بعدي)
        pre_exam = TeacherExam.objects.create(
            skill=skill, exam_type='pre',
            questions_count=pre_count, duration_minutes=pre_time,
            pass_score=pre_pass, delivery=request.POST.get('pre_delivery', 'both'),
            is_active=is_active,
        )
        qs_saved += _bulk_questions(pre_exam, request.POST.get('pre_questions'), 'pre')

        post_exam = TeacherExam.objects.create(
            skill=skill, exam_type='post',
            questions_count=post_count, duration_minutes=post_time,
            pass_score=post_pass, is_active=is_active,
        )
        qs_saved += _bulk_questions(post_exam, request.POST.get('post_questions'), 'post')
    else:
        # درس تحصيلي / بنك → اختبار واحد
        exam = TeacherExam.objects.create(
            skill=skill,
            exam_type='lesson' if content_type == 'lesson' else 'bank',
            questions_count=pre_count, duration_minutes=pre_time,
            pass_score=pre_pass, is_active=is_active,
        )
        qs_saved += _bulk_questions(exam, request.POST.get('pre_questions'), 'pre')

    # بعد الحفظ — نوجّه المعلمة للمحرر المتقدم (مثل صفحة المديرة) لإكمال الأسئلة
    # بصور وLaTeX و MathLive و OCR، إن لم يكن قد كملت إدخالها في الـ wizard.
    first_exam = TeacherExam.objects.filter(skill=skill).order_by('id').first()
    if qs_saved:
        messages.success(
            request,
            f'✅ تم حفظ "{title}" مع {qs_saved} سؤال — '
            'يمكنك إكمال الأسئلة بالمحرر المتقدم (LaTeX + صور + مسح ضوئي)'
        )
    else:
        messages.success(
            request,
            f'✅ تم حفظ "{title}" — افتحي المحرر المتقدم لإضافة الأسئلة بكل الميزات'
        )

    # إن طلبت المعلمة الانتقال المباشر للمحرر، نحوّلها
    if first_exam and request.POST.get('open_advanced') == '1':
        return redirect('add_question', exam_id=first_exam.id)
    return redirect('skill_manager')


@login_required
def delete_skill(request, skill_id):
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    skill = get_object_or_404(TeacherSkill, id=skill_id, created_by=teacher)
    skill.delete()
    messages.success(request, 'تم حذف المهارة')
    return redirect('skill_manager')


@login_required
def use_shared_skill(request, skill_id):
    """نسخ مهارة من زميلة لاستخدامها في فصلي."""
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    original = get_object_or_404(TeacherSkill, id=skill_id, is_shared=True)

    if request.method == 'POST':
        # نسخ المهارة
        new_skill = TeacherSkill.objects.create(
            created_by=teacher,
            title=original.title,
            description=original.description,
            content_type=original.content_type,
            skill_type=original.skill_type,
            subject=original.subject,
            target_classes=original.target_classes,
            is_shared=False,
            is_active=False,
        )
        # نسخ المحتوى إن وجد
        if hasattr(original, 'content') and original.content:
            TeacherSkillContent.objects.create(
                skill=new_skill,
                plain_text=original.content.plain_text or '',
                video_url=original.content.video_url or '',
            )
        # نسخ الاختبارات والأسئلة
        for exam in original.exams.all():
            new_exam = TeacherExam.objects.create(
                skill=new_skill,
                exam_type=exam.exam_type,
                questions_count=exam.questions_count,
                duration_minutes=exam.duration_minutes,
                pass_score=exam.pass_score,
                delivery=exam.delivery,
                is_active=False,
            )
            for q in exam.questions.all():
                TeacherQuestion.objects.create(
                    exam=new_exam,
                    order=q.order,
                    question_plain=q.question_plain,
                    option_a_plain=q.option_a_plain,
                    option_b_plain=q.option_b_plain,
                    option_c_plain=q.option_c_plain,
                    option_d_plain=q.option_d_plain,
                    correct_answer=q.correct_answer,
                    target_skill_name=q.target_skill_name,
                    feedback_plain=q.feedback_plain,
                )
        messages.success(request, f'✅ تم نسخ المهارة "{original.title}" — يمكنكِ الآن تعديلها وتفعيلها')
        return redirect('skill_manager')

    return redirect('skill_manager')


@login_required
def add_question(request, exam_id):
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    exam = get_object_or_404(TeacherExam, id=exam_id, skill__created_by=teacher)
    if request.method == 'POST':
        # ──────── تشخيص شامل ────────
        import logging as _log
        _lg = _log.getLogger('teachers.add_question')
        _lg.warning(f"📥 add_question POST — exam_id={exam_id}")
        _lg.warning(f"   Content-Type: {request.content_type}")
        _lg.warning(f"   FILES keys: {list(request.FILES.keys())}")
        _lg.warning(f"   POST keys: {list(request.POST.keys())}")
        for k, f in request.FILES.items():
            _lg.warning(f"   FILE [{k}]: name={f.name}, size={f.size}, type={f.content_type}")
        # ────────────────────────────

        order = exam.questions.count() + 1
        # ربط المهارة المعيارية (إن اختارت)
        skill_standard_id = request.POST.get('skill_standard', '')
        skill_standard_obj = None
        if skill_standard_id:
            try:
                skill_standard_obj = SkillStandard.objects.get(id=int(skill_standard_id))
            except (SkillStandard.DoesNotExist, ValueError):
                pass

        q = TeacherQuestion.objects.create(
            exam=exam, order=order,
            question_plain=request.POST.get('question_plain', '').strip(),
            option_a_plain=request.POST.get('option_a_plain', ''),
            option_b_plain=request.POST.get('option_b_plain', ''),
            option_c_plain=request.POST.get('option_c_plain', ''),
            option_d_plain=request.POST.get('option_d_plain', ''),
            correct_answer=request.POST.get('correct_answer', 'A'),
            target_skill_name=request.POST.get('target_skill_name', ''),
            skill_standard=skill_standard_obj,
            feedback_plain=request.POST.get('feedback_plain', ''),
        )
        # دعم الصور للسؤال + الخيارات (يقبل ملف أو data:URL من OCR)
        from core.views import _attach_image as _att
        upload_results = {}
        for fld in ('question_image', 'option_a_image', 'option_b_image', 'option_c_image', 'option_d_image'):
            result = _att(q, fld, request)
            if result is not None:
                upload_results[fld] = result
            _lg.warning(f"   _attach_image({fld}) → {result}")
        q.save()
        # عرض نتائج الرفع للمستخدم
        failed = [k for k, v in upload_results.items() if v is False]
        if failed:
            messages.warning(request, f'⚠️ فشل رفع بعض الصور: {", ".join(failed)}')

        # رسالة تشخيصية مرئية للمستخدم
        file_keys = list(request.FILES.keys())
        succeeded = [k for k, v in upload_results.items() if v is True]
        if file_keys:
            messages.info(request, f'📎 ملفات وصلت: {", ".join(file_keys)} | نجح: {", ".join(succeeded) if succeeded else "لا شيء"}')
        elif not upload_results:
            messages.info(request, '📎 لم تصل أي ملفات مع الطلب (request.FILES فارغ)')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'order': order})
        messages.success(request, '✅ تم إضافة السؤال')
        return redirect('add_question', exam_id=exam.id)

    # قائمة مهارات كل المدرسة (لربط أسئلة البنوك بأي مهارة لأي معلمة)
    school_skills = (
        TeacherSkill.objects
        .exclude(content_type='bank')
        .filter(is_active=True)
        .select_related('created_by')
        .order_by('created_by__full_name', 'title')
    )

    # المهارات المعيارية لعرضها في القائمة المنسدلة
    skill_standards = SkillStandard.objects.filter(is_active=True).order_by('track', 'order')

    return render(request, 'teachers/add_question.html', {
        'exam': exam, 'teacher': teacher,
        'school_skills': school_skills,
        'skill_standards': skill_standards,
        'questions': exam.questions.order_by('order'),
    })


@login_required
def import_questions_excel(request, exam_id):
    try:
        is_admin = request.user.core_profile.role == 'ADMIN'
    except:
        is_admin = False

    if not is_admin and not check_teacher(request):
        return redirect('home')

    if is_admin:
        exam = get_object_or_404(TeacherExam, id=exam_id)
    else:
        teacher = get_teacher(request)
        exam = get_object_or_404(TeacherExam, id=exam_id, skill__created_by=teacher)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                TeacherQuestion.objects.create(
                    exam=exam, order=exam.questions.count()+1,
                    question_plain=str(row[0]) if row[0] else '',
                    option_a_plain=str(row[1]) if row[1] else '',
                    option_b_plain=str(row[2]) if row[2] else '',
                    option_c_plain=str(row[3]) if row[3] else '',
                    option_d_plain=str(row[4]) if row[4] else '',
                    correct_answer=str(row[5]).upper() if row[5] else 'A',
                    target_skill_name=str(row[6]) if len(row)>6 and row[6] else '',
                    feedback_plain=str(row[7]) if len(row)>7 and row[7] else '',
                )
                count += 1
            messages.success(request, f'تم استيراد {count} سؤال!')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')

    if is_admin:
        return redirect('admin_comprehensive')
    return redirect('skill_manager')


@login_required
def import_skills_excel(request):
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    if not teacher:
        return redirect('teacher_dashboard')
    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                title = str(row[1]).strip() if row[1] else ''
                if not title: continue
                TeacherSkill.objects.create(
                    content_type=str(row[0]).strip() if row[0] else 'skill',
                    title=title,
                    skill_type=str(row[2]).strip() if len(row)>2 and row[2] else '',
                    subject=str(row[3]).strip() if len(row)>3 and row[3] else '',
                    target_classes=str(row[4]).strip() if len(row)>4 and row[4] else '',
                    description=str(row[5]).strip() if len(row)>5 and row[5] else '',
                    created_by=teacher, is_active=True,
                )
                count += 1
            messages.success(request, f'تم استيراد {count} مهارة!')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    return redirect('skill_manager')


@login_required
def exam_results(request, exam_id):
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    exam = get_object_or_404(TeacherExam, id=exam_id, skill__created_by=teacher)
    results = ExamResult.objects.filter(exam=exam).select_related('student', 'student_record').order_by('-percentage')
    avg = results.aggregate(avg=Avg('percentage'))['avg'] or 0
    passed = results.filter(passed=True).count()
    context = {
        'exam': exam,
        'results': results,
        'avg': int(round(avg)),
        'passed': passed,
        'failed': results.count()-passed,
        'teacher': teacher,
    }
    return render(request, 'teachers/exam_results.html', context)


@login_required
def student_result(request, result_id):
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    result = get_object_or_404(ExamResult, id=result_id, exam__skill__created_by=teacher)
    answers = StudentAnswer.objects.filter(result=result).select_related('question').order_by('question__order')
    return render(request, 'teachers/student_result.html', {'result': result, 'answers': answers, 'teacher': teacher})


@login_required
def enter_manual_score(request, result_id):
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    result = get_object_or_404(ExamResult, id=result_id, exam__skill__created_by=teacher)
    if request.method == 'POST':
        score = float(request.POST.get('score', 0))
        result.score = score
        result.percentage = (score / result.total * 100) if result.total > 0 else 0
        result.passed = result.percentage >= result.exam.pass_score
        result.manually_corrected = True
        result.corrected_by = teacher
        result.save()
        messages.success(request, f'تم رصد الدرجة: {score}/{result.total}')
        return redirect('student_result', result_id=result_id)
    return render(request, 'teachers/enter_score.html', {'result': result, 'teacher': teacher})


@login_required
def add_session(request):
    """تسجيل حصة تدريبية يدوية — تحفظ حق المعلمة في الإحصاء."""
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    if not teacher:
        messages.error(request, 'لا يوجد سجل معلمة لحسابك')
        return redirect('teacher_dashboard')

    if request.method == 'POST':
        skill_id = (request.POST.get('skill_id') or '').strip()
        skill = TeacherSkill.objects.filter(id=skill_id).first() if skill_id else None
        if skill is None:
            messages.error(request, 'يرجى اختيار المهارة/الدرس')
            return redirect('add_session')

        # نوع الحصة يُستنتج من نوع المحتوى تلقائياً (يمكن التعديل يدوياً)
        session_type = request.POST.get('session_type', '').strip()
        if not session_type:
            session_type = ('qodrat' if skill.skill_type in ('qodrat_kamy', 'qodrat_lafzy')
                            else 'tahsili' if skill.subject in ('math', 'bio', 'chem', 'phys')
                            else 'qodrat')

        from datetime import date as _date, time as _time
        # دعم اختيار فصول متعددة
        target_classes = request.POST.getlist('target_class')
        if not target_classes:
            target_classes = ['']   # سجّل حصة بدون فصل محدد
        session_date = request.POST.get('session_date') or _date.today().isoformat()
        session_time = request.POST.get('session_time') or _time(8, 0).isoformat()
        notes        = request.POST.get('notes', '')
        try:
            for cls in target_classes:
                ClassSession.objects.create(
                    teacher=teacher, skill=skill,
                    session_type=session_type,
                    target_class=cls,
                    session_date=session_date,
                    session_time=session_time,
                    notes=notes,
                )
            count = len(target_classes)
            messages.success(request, f'✅ تم تسجيل {"الحصة" if count == 1 else f"{count} حصص"} وستُحتسب في إحصائك')
        except Exception as exc:
            messages.error(request, f'تعذّر التسجيل: {exc}')
            return redirect('add_session')

        return redirect('teacher_dashboard')

    # كل مهارات المعلمة (مشاركة + خاصة + بنوك) لكي تختار من ضمنها
    my_skills = (TeacherSkill.objects.filter(created_by=teacher).order_by('-created_at')
                 if teacher else [])
    classrooms = teacher.classrooms.all().order_by('name')
    return render(request, 'teachers/add_session.html', {
        'teacher': teacher, 'skills': my_skills, 'classrooms': classrooms,
    })


# ═══════════════════════════════════════════════════════════════
# عرض / تعديل / تعديل أسئلة (CRUD كامل للمعلمة)
# ═══════════════════════════════════════════════════════════════

@login_required
def view_skill(request, skill_id):
    """عرض المهارة كاملة (محتوى + أسئلة) — للمعلمة المنشئة أو المشاركة."""
    if not check_teacher_or_admin(request):
        return redirect('home')
    teacher = get_teacher(request)
    skill = get_object_or_404(TeacherSkill, id=skill_id)
    # الصلاحية: المنشئة، أو المهارة مشتركة، أو المديرة
    is_owner = teacher and skill.created_by_id == teacher.id
    is_admin = False
    try:
        is_admin = request.user.core_profile.role == 'ADMIN'
    except Profile.DoesNotExist:
        pass
    if not (is_owner or skill.is_shared or is_admin):
        messages.error(request, 'لا تملكين صلاحية عرض هذه المهارة')
        return redirect('skill_manager')

    return render(request, 'teachers/view_skill.html', {
        'skill': skill,
        'content': getattr(skill, 'content', None),
        'exams': skill.exams.all().prefetch_related('questions'),
        'is_owner': is_owner or is_admin,
        'teacher': teacher,
    })


@login_required
def serve_skill_pdf(request, content_id):
    """Proxy: يجلب PDF من Cloudinary ويقدّمه للمتصفح من نفس الأصل لتجنب CORS."""
    from django.http import HttpResponse, Http404
    import urllib.request as _urlreq
    content = get_object_or_404(TeacherSkillContent, id=content_id)
    if not content.pdf_file:
        raise Http404
    try:
        import cloudinary.utils
        pdf_url, _ = cloudinary.utils.cloudinary_url(str(content.pdf_file), resource_type='raw')
        with _urlreq.urlopen(pdf_url, timeout=30) as resp:
            data = resp.read()
        return HttpResponse(data, content_type='application/pdf')
    except Exception:
        raise Http404


@login_required
def edit_skill(request, skill_id):
    """تعديل بيانات المهارة + المحتوى + إعداداتها."""
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    skill = get_object_or_404(TeacherSkill, id=skill_id, created_by=teacher)

    if request.method == 'POST':
        skill.title = (request.POST.get('title') or skill.title).strip()
        skill.description = request.POST.get('description', skill.description)
        skill.target_classes = request.POST.get('target_classes', skill.target_classes)
        skill.is_shared = request.POST.get('is_shared') == 'on'
        skill.is_active = request.POST.get('is_active') == 'on'
        skill.skill_type = request.POST.get('skill_type', skill.skill_type)
        skill.subject = request.POST.get('subject', skill.subject)
        skill.save()

        # تحديث/إنشاء المحتوى
        plain_text = request.POST.get('plain_text', '').strip()
        video_url = request.POST.get('video_url', '').strip()
        content, _ = TeacherSkillContent.objects.get_or_create(skill=skill)
        if plain_text:
            content.plain_text = plain_text
        if video_url:
            content.video_url = video_url
        if 'plain_image' in request.FILES:
            try:
                import cloudinary.uploader
                r = cloudinary.uploader.upload(request.FILES['plain_image'], folder='skill_content')
                content.plain_image = r['public_id']
            except Exception:
                pass
        if 'pdf_file' in request.FILES:
            pdf_f = request.FILES['pdf_file']
            if pdf_f.size > 10 * 1024 * 1024:
                messages.warning(request, f'⚠️ ملف PDF كبير ({round(pdf_f.size/1048576,1)} ميجا). الحد الأقصى 10 ميجا.')
            else:
                try:
                    import cloudinary.uploader
                    r = cloudinary.uploader.upload(pdf_f, folder='skill_pdfs', resource_type='raw', access_mode='public')
                    content.pdf_file = r['public_id']
                except Exception as e:
                    messages.warning(request, f'⚠️ فشل رفع PDF: {e}')
        content.save()

        messages.success(request, '✅ تم حفظ التعديلات')
        return redirect('skill_manager')

    return render(request, 'teachers/edit_skill.html', {
        'skill': skill,
        'content': getattr(skill, 'content', None),
        'teacher': teacher,
    })


@login_required
def edit_question(request, exam_id, q_id):
    """تعديل سؤال داخل اختبار."""
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    exam = get_object_or_404(TeacherExam, id=exam_id, skill__created_by=teacher)
    q = get_object_or_404(TeacherQuestion, id=q_id, exam=exam)

    if request.method == 'POST':
        q.question_plain = request.POST.get('question_plain', q.question_plain).strip()
        q.option_a_plain = request.POST.get('option_a_plain', q.option_a_plain)
        q.option_b_plain = request.POST.get('option_b_plain', q.option_b_plain)
        q.option_c_plain = request.POST.get('option_c_plain', q.option_c_plain)
        q.option_d_plain = request.POST.get('option_d_plain', q.option_d_plain)
        q.correct_answer = request.POST.get('correct_answer', q.correct_answer).upper()[:1]
        q.target_skill_name = request.POST.get('target_skill_name', q.target_skill_name)
        q.feedback_plain = request.POST.get('feedback_plain', q.feedback_plain)
        # ربط المهارة المعيارية
        skill_standard_id = request.POST.get('skill_standard', '')
        if skill_standard_id:
            try:
                q.skill_standard = SkillStandard.objects.get(id=int(skill_standard_id))
            except (SkillStandard.DoesNotExist, ValueError):
                pass
        else:
            q.skill_standard = None
        # دعم الصور للسؤال + الخيارات (ملف أو data:URL)
        from core.views import _attach_image as _att
        upload_results = {}
        for fld in ('question_image', 'option_a_image', 'option_b_image', 'option_c_image', 'option_d_image'):
            result = _att(q, fld, request)
            if result is not None:
                upload_results[fld] = result
        q.save()
        failed = [k for k, v in upload_results.items() if v is False]
        if failed:
            messages.warning(request, f'⚠️ فشل رفع بعض الصور: {", ".join(failed)}')
        messages.success(request, '✅ تم حفظ السؤال')
        return redirect('view_skill', skill_id=exam.skill_id)

    # المهارات المعيارية لعرضها في القائمة المنسدلة
    skill_standards = SkillStandard.objects.filter(is_active=True).order_by('track', 'order')

    return render(request, 'teachers/edit_question.html', {
        'q': q, 'exam': exam, 'teacher': teacher,
        'skill_standards': skill_standards,
    })


@login_required
def delete_question(request, exam_id, q_id):
    """حذف سؤال."""
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    exam = get_object_or_404(TeacherExam, id=exam_id, skill__created_by=teacher)
    q = get_object_or_404(TeacherQuestion, id=q_id, exam=exam)
    if request.method == 'POST':
        q.delete()
        messages.success(request, '🗑️ تم حذف السؤال')
    return redirect('view_skill', skill_id=exam.skill_id)


# ═══════════════════════════════════════════════════════════════

@login_required
def get_skill_questions(request, skill_id):
    teacher = get_teacher(request)
    skill = get_object_or_404(TeacherSkill, id=skill_id, created_by=teacher)
    data = {'skill': skill.title, 'exams': []}
    for exam in skill.exams.all():
        exam_data = {
            'type': exam.exam_type,
            'type_label': exam.get_exam_type_display(),
            'count': exam.questions.count(),
            'time': exam.duration_minutes,
            'pass_score': exam.pass_score,
            'questions': []
        }
        for q in exam.questions.all().order_by('order'):
            exam_data['questions'].append({
                'order': q.order,
                'text': q.question_plain,
                'a': q.option_a_plain,
                'b': q.option_b_plain,
                'c': q.option_c_plain,
                'd': q.option_d_plain,
                'correct': q.correct_answer,
                'feedback': q.feedback_plain,
            })
        data['exams'].append(exam_data)
    return JsonResponse(data)


# ═══════════════════════════════════════════════════════════════
# إدارة فصول المعلمة + الرصد اليدوي
# ═══════════════════════════════════════════════════════════════

@login_required
def manage_classrooms(request):
    """إدارة فصول المعلمة — ربط/فك ربط فصول."""
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)
    if request.method == 'POST':
        action = request.POST.get('action')
        classroom_id = request.POST.get('classroom_id')
        if action == 'add' and classroom_id:
            cr = ClassRoom.objects.filter(id=classroom_id).first()
            if cr:
                teacher.classrooms.add(cr)
                messages.success(request, f'تم إضافة الفصل: {cr.name}')
        elif action == 'remove' and classroom_id:
            cr = ClassRoom.objects.filter(id=classroom_id).first()
            if cr:
                teacher.classrooms.remove(cr)
                messages.success(request, f'تم إزالة الفصل: {cr.name}')
        return redirect('manage_classrooms')

    my_classrooms = teacher.classrooms.all().prefetch_related('students')
    available = ClassRoom.objects.exclude(id__in=my_classrooms.values_list('id', flat=True))
    return render(request, 'teachers/manage_classrooms.html', {
        'teacher': teacher,
        'my_classrooms': my_classrooms,
        'available_classrooms': available,
    })


@login_required
def get_classroom_students(request, classroom_id):
    """API — جلب طالبات فصل معيّن (JSON)."""
    teacher = get_teacher(request)
    if not teacher:
        return JsonResponse({'error': 'unauthorized'}, status=403)
    cr = get_object_or_404(ClassRoom, id=classroom_id)
    students = cr.students.all().order_by('full_name')
    return JsonResponse({
        'students': [{'id': s.id, 'name': s.full_name} for s in students]
    })


@login_required
def skill_standards_json(request):
    """يرجع المهارات المعيارية — يمكن فلترتها بالمسار."""
    track = request.GET.get('track', '')
    show_all = request.GET.get('all', '')  # لعرض المعطّلة أيضاً في صفحة الإدارة
    if show_all:
        qs = SkillStandard.objects.all()
    else:
        qs = SkillStandard.objects.filter(is_active=True)
    if track:
        qs = qs.filter(track=track)
    data = [{'id': s.id, 'code': s.code, 'name': s.name, 'track': s.track,
             'track_label': s.track_label, 'description': s.description,
             'is_active': s.is_active}
            for s in qs.order_by('track', 'order')]
    return JsonResponse({'skills': data})


@login_required
def manage_skill_standard(request):
    """إضافة أو تعديل مهارة معيارية (POST JSON)."""
    import json
    if not check_teacher_or_admin(request):
        return JsonResponse({'error': 'غير مصرّح'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'POST مطلوب'}, status=405)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'بيانات غير صالحة'}, status=400)

    action = body.get('action', 'add')  # add | edit | toggle | delete
    skill_id = body.get('id')

    if action == 'add':
        code = (body.get('code') or '').strip()
        name = (body.get('name') or '').strip()
        track = body.get('track', '')
        if not code or not name or not track:
            return JsonResponse({'error': 'الرمز والاسم والمسار مطلوبة'}, status=400)
        if SkillStandard.objects.filter(code=code).exists():
            return JsonResponse({'error': f'الرمز {code} موجود مسبقاً'}, status=400)
        # حساب الترتيب التالي في المسار
        max_order = SkillStandard.objects.filter(track=track).aggregate(
            m=models.Max('order'))['m'] or 0
        ss = SkillStandard.objects.create(
            code=code, name=name, track=track,
            description=body.get('description', ''),
            order=max_order + 1, is_active=True,
        )
        return JsonResponse({'ok': True, 'id': ss.id, 'message': f'تم إضافة "{name}"'})

    elif action == 'edit':
        if not skill_id:
            return JsonResponse({'error': 'معرّف المهارة مطلوب'}, status=400)
        try:
            ss = SkillStandard.objects.get(id=skill_id)
        except SkillStandard.DoesNotExist:
            return JsonResponse({'error': 'المهارة غير موجودة'}, status=404)
        if body.get('name'):
            ss.name = body['name'].strip()
        if body.get('description') is not None:
            ss.description = body['description']
        if body.get('code'):
            new_code = body['code'].strip()
            if new_code != ss.code and SkillStandard.objects.filter(code=new_code).exists():
                return JsonResponse({'error': f'الرمز {new_code} موجود مسبقاً'}, status=400)
            ss.code = new_code
        if body.get('track'):
            ss.track = body['track']
        ss.save()
        return JsonResponse({'ok': True, 'message': f'تم تعديل "{ss.name}"'})

    elif action == 'toggle':
        if not skill_id:
            return JsonResponse({'error': 'معرّف المهارة مطلوب'}, status=400)
        try:
            ss = SkillStandard.objects.get(id=skill_id)
        except SkillStandard.DoesNotExist:
            return JsonResponse({'error': 'المهارة غير موجودة'}, status=404)
        ss.is_active = not ss.is_active
        ss.save()
        state = 'مفعّلة' if ss.is_active else 'معطّلة'
        return JsonResponse({'ok': True, 'is_active': ss.is_active,
                             'message': f'"{ss.name}" الآن {state}'})

    elif action == 'delete':
        if not skill_id:
            return JsonResponse({'error': 'معرّف المهارة مطلوب'}, status=400)
        try:
            ss = SkillStandard.objects.get(id=skill_id)
        except SkillStandard.DoesNotExist:
            return JsonResponse({'error': 'المهارة غير موجودة'}, status=404)
        # تحقق أن المهارة ما عليها أسئلة مرتبطة
        q_count = ss.questions.count()
        if q_count > 0:
            return JsonResponse({
                'error': f'لا يمكن حذف المهارة — مرتبطة بـ {q_count} سؤال. عطّليها بدل الحذف.',
            }, status=400)
        name = ss.name
        ss.delete()
        return JsonResponse({'ok': True, 'message': f'تم حذف "{name}"'})

    return JsonResponse({'error': 'إجراء غير معروف'}, status=400)


@login_required
def gap_analysis_json(request):
    """
    محرك تحليل الفجوات — يحسب نسبة الإتقان لكل مهارة معيارية.
    GET params:
      scope  = student | classroom | school
      key    = u_123 | sr_456 | classroom_id (حسب scope)
      track  = qodrat_kamy | qodrat_lafzy | tahsili_math | ... (اختياري)
    يرجع: { skills: [{code, name, track, track_label, total, correct, mastery, status}], summary: {...} }
    """
    if not check_teacher_or_admin(request):
        return JsonResponse({'error': 'غير مصرّح'}, status=403)

    scope = request.GET.get('scope', 'school')
    key = request.GET.get('key', '')
    track_filter = request.GET.get('track', '')

    teacher = get_teacher(request)

    # بناء queryset الإجابات المرتبطة بمعلمة (مفلترة للمعلمة الحالية فقط)
    answers_qs = StudentAnswer.objects.filter(
        question__skill_standard__isnull=False,
        result__exam__skill__created_by__user=request.user
    )

    # تحديد النطاق
    if scope == 'student':
        if key.startswith('u_'):
            try:
                uid = int(key[2:])
                answers_qs = answers_qs.filter(result__student_id=uid)
            except ValueError:
                return JsonResponse({'error': 'مفتاح غير صالح'}, status=400)
        elif key.startswith('sr_'):
            try:
                sr_id = int(key[3:])
                answers_qs = answers_qs.filter(result__student_record_id=sr_id)
            except ValueError:
                return JsonResponse({'error': 'مفتاح غير صالح'}, status=400)
        else:
            return JsonResponse({'error': 'مفتاح طالبة مطلوب'}, status=400)
    elif scope == 'classroom':
        if key:
            try:
                classroom_id = int(key)
                classroom = ClassRoom.objects.get(id=classroom_id)
                student_ids = classroom.students.values_list('user_id', flat=True)
                sr_ids = classroom.students.values_list('id', flat=True)
                answers_qs = answers_qs.filter(
                    Q(result__student_id__in=student_ids) | Q(result__student_record_id__in=sr_ids)
                )
            except (ValueError, ClassRoom.DoesNotExist):
                return JsonResponse({'error': 'فصل غير موجود'}, status=400)
    # scope == 'school' → no extra filter

    # فلترة بالمسار
    if track_filter:
        answers_qs = answers_qs.filter(question__skill_standard__track=track_filter)

    # تجميع النتائج حسب المهارة المعيارية
    from django.db.models import Sum, Case, When, IntegerField, Value
    skill_stats = (
        answers_qs
        .values(
            'question__skill_standard__id',
            'question__skill_standard__code',
            'question__skill_standard__name',
            'question__skill_standard__track',
        )
        .annotate(
            total=Count('id'),
            correct=Sum(Case(When(is_correct=True, then=Value(1)), default=Value(0), output_field=IntegerField())),
        )
        .order_by('question__skill_standard__track', 'question__skill_standard__order')
    )

    # بناء الاستجابة
    TRACK_LABELS = dict(SkillStandard.TRACK_CHOICES)
    skills_data = []
    total_q = 0
    total_correct = 0
    mastered = 0  # >= 70%
    developing = 0  # 40-69%
    weak = 0  # < 40%

    for s in skill_stats:
        t = s['total'] or 0
        c = s['correct'] or 0
        mastery = round((c / t * 100) if t > 0 else 0, 1)
        trk = s['question__skill_standard__track'] or ''

        if mastery >= 70:
            status = 'mastered'
            mastered += 1
        elif mastery >= 40:
            status = 'developing'
            developing += 1
        else:
            status = 'weak'
            weak += 1

        total_q += t
        total_correct += c

        skills_data.append({
            'id': s['question__skill_standard__id'],
            'code': s['question__skill_standard__code'],
            'name': s['question__skill_standard__name'],
            'track': trk,
            'track_label': TRACK_LABELS.get(trk, trk),
            'total': t,
            'correct': c,
            'mastery': mastery,
            'status': status,
        })

    # ترتيب: الأضعف أولاً (لتحديد الفجوات)
    skills_data.sort(key=lambda x: x['mastery'])

    # المهارات المعيارية التي لا يوجد لها إجابات (فجوات غير مُختبرة)
    tested_ids = {s['id'] for s in skills_data}
    untested_qs = SkillStandard.objects.filter(
        is_active=True,
        questions__exam__skill__created_by__user=request.user
    ).distinct().exclude(id__in=tested_ids)
    if track_filter:
        untested_qs = untested_qs.filter(track=track_filter)
    untested = [
        {
            'id': s.id, 'code': s.code, 'name': s.name,
            'track': s.track, 'track_label': s.track_label,
            'total': 0, 'correct': 0, 'mastery': 0, 'status': 'untested',
        }
        for s in untested_qs.order_by('track', 'order')
    ]

    overall_mastery = round((total_correct / total_q * 100) if total_q > 0 else 0, 1)

    # توصيات
    recommendations = []
    weak_skills = [s for s in skills_data if s['status'] == 'weak']
    if weak_skills:
        recommendations.append(f"يوجد {len(weak_skills)} مهارة ضعيفة تحتاج تركيز فوري")
        for ws in weak_skills[:5]:
            recommendations.append(f"• [{ws['code']}] {ws['name']} — نسبة الإتقان: {ws['mastery']}%")
    if untested:
        recommendations.append(f"يوجد {len(untested)} مهارة لم تُختبر بعد")

    return JsonResponse({
        'skills': skills_data,
        'untested': untested,
        'summary': {
            'total_questions': total_q,
            'total_correct': total_correct,
            'overall_mastery': overall_mastery,
            'mastered_count': mastered,
            'developing_count': developing,
            'weak_count': weak,
            'untested_count': len(untested),
            'total_standards': len(skills_data) + len(untested),
        },
        'recommendations': recommendations,
    })


@login_required
def teacher_student_report_json(request):
    """يرجع بيانات تقرير طالبة معيّنة — تحليل مهارات + سجل اختبارات."""
    if not check_teacher(request):
        return JsonResponse({'error': 'غير مصرّح'}, status=403)

    teacher = get_teacher(request)
    key = request.GET.get('key', '')  # u_123 or sr_456

    if not key:
        return JsonResponse({'error': 'مفتاح الطالبة مطلوب'}, status=400)

    # بناء queryset النتائج حسب نوع المفتاح
    my_results = ExamResult.objects.filter(exam__skill__created_by=teacher)

    if key.startswith('u_'):
        try:
            student_id = int(key[2:])
        except ValueError:
            return JsonResponse({'error': 'مفتاح غير صالح'}, status=400)
        qs = my_results.filter(student_id=student_id).select_related('exam', 'exam__skill')
    elif key.startswith('sr_'):
        try:
            sr_id = int(key[3:])
        except ValueError:
            return JsonResponse({'error': 'مفتاح غير صالح'}, status=400)
        qs = my_results.filter(student_record_id=sr_id).select_related('exam', 'exam__skill')
    else:
        return JsonResponse({'error': 'مفتاح غير معروف'}, status=400)

    # ── سجل الاختبارات ──
    exams_list = []
    for r in qs.order_by('-submitted_at'):
        exams_list.append({
            'exam_name': r.exam.skill.title if r.exam.skill else (r.exam.title or '—'),
            'exam_type': r.exam.get_exam_type_display() if hasattr(r.exam, 'get_exam_type_display') else r.exam.exam_type,
            'percentage': float(r.percentage or 0),
            'passed': r.passed,
            'date': r.submitted_at.strftime('%Y-%m-%d') if r.submitted_at else '—',
            'score': f'{r.score or 0}/{r.total or 0}',
        })

    # ── تحليل المهارات ──
    from collections import defaultdict
    skill_bucket = defaultdict(lambda: {'correct': 0, 'total': 0})
    for ans in StudentAnswer.objects.filter(result__in=qs).select_related('question'):
        name = (ans.question.target_skill_name or 'غير محدّدة').strip() or 'غير محدّدة'
        skill_bucket[name]['total'] += 1
        if ans.is_correct:
            skill_bucket[name]['correct'] += 1

    skills = []
    mastered = 0
    needs = 0
    for name, v in skill_bucket.items():
        if v['total'] == 0:
            continue
        pct = round(100 * v['correct'] / v['total'])
        skills.append({'name': name, 'pct': pct, 'correct': v['correct'], 'total': v['total']})
        if pct >= 70:
            mastered += 1
        else:
            needs += 1
    skills.sort(key=lambda r: -r['pct'])

    # ── مقارنة قبلي/بعدي ──
    pre_post = {}
    for r in qs:
        sk_name = r.exam.skill.title if r.exam.skill else '—'
        if sk_name not in pre_post:
            pre_post[sk_name] = {'pre': None, 'post': None}
        if r.exam.exam_type == 'pre':
            pre_post[sk_name]['pre'] = float(r.percentage or 0)
        elif r.exam.exam_type == 'post':
            pre_post[sk_name]['post'] = float(r.percentage or 0)
    pre_post_list = [
        {'skill': k, 'pre': v['pre'], 'post': v['post'],
         'improvement': round(v['post'] - v['pre'], 1) if v['pre'] is not None and v['post'] is not None else None}
        for k, v in pre_post.items()
        if v['pre'] is not None or v['post'] is not None
    ]

    return JsonResponse({
        'exams': exams_list,
        'skills': skills,
        'mastered': mastered,
        'needs': needs,
        'pre_post': pre_post_list,
        'total_exams': len(exams_list),
    })


@login_required
def manual_score_entry(request):
    """رصد درجة يدوي — المعلمة تختار الاختبار والطالبة وتدخل الدرجة."""
    if not check_teacher(request):
        return redirect('home')
    teacher = get_teacher(request)

    if request.method == 'POST':
        exam_id = request.POST.get('exam_id')
        student_id = request.POST.get('student_id')
        score = request.POST.get('score', '0')
        exam = get_object_or_404(TeacherExam, id=exam_id, skill__created_by=teacher)
        student_obj = get_object_or_404(Student, id=student_id)

        try:
            score = float(score)
        except (ValueError, TypeError):
            score = 0

        total = exam.questions_count or exam.questions.count() or 10
        percentage = (score / total * 100) if total > 0 else 0
        passed = percentage >= exam.pass_score

        # تحقق من عدم التكرار
        existing = ExamResult.objects.filter(exam=exam, student_record=student_obj).first()
        if existing:
            existing.score = score
            existing.total = total
            existing.percentage = percentage
            existing.passed = passed
            existing.manually_corrected = True
            existing.corrected_by = teacher
            existing.save()
            messages.success(request, f'تم تحديث درجة {student_obj.full_name}: {score}/{total}')
        else:
            ExamResult.objects.create(
                exam=exam,
                student_record=student_obj,
                score=score,
                total=total,
                percentage=percentage,
                passed=passed,
                manually_corrected=True,
                corrected_by=teacher,
            )
            messages.success(request, f'تم رصد درجة {student_obj.full_name}: {score}/{total}')

        # ارجع للصفحة اللي جاء منها
        next_url = request.POST.get('next', '')
        if next_url:
            return redirect(next_url)
        return redirect('teacher_dashboard')

    # GET — عرض الفورم
    my_classrooms = teacher.classrooms.all().prefetch_related('students')
    my_exams = TeacherExam.objects.filter(skill__created_by=teacher).select_related('skill')
    return render(request, 'teachers/manual_score.html', {
        'teacher': teacher,
        'classrooms': my_classrooms,
        'exams': my_exams,
    })


# ═══════════════════════════════════════════════════════════════
# API: إحصاءات الداشبورد — للتحديث التلقائي (AJAX)
# ═══════════════════════════════════════════════════════════════
@login_required
def teacher_exams_json(request):
    """API — جلب اختبارات المعلمة المفعّلة لتعيينها علاجياً."""
    if not check_teacher(request):
        return JsonResponse({'error': 'غير مصرّح'}, status=403)
    teacher = get_teacher(request)
    if not teacher:
        return JsonResponse({'exams': []})
    exams = (
        TeacherExam.objects
        .filter(skill__created_by=teacher, is_active=True)
        .select_related('skill')
        .order_by('skill__title', 'exam_type')
    )
    data = []
    for e in exams:
        data.append({
            'id': e.id,
            'skill': e.skill.title if e.skill else '—',
            'type': e.get_exam_type_display() if hasattr(e, 'get_exam_type_display') else e.exam_type,
            'questions': e.questions_count,
        })
    return JsonResponse({'exams': data})


@login_required
def assign_remedial_exam(request):
    """API — تعيين اختبار علاجي لطالبة محددة مباشرة."""
    import json as _json
    if not check_teacher(request):
        return JsonResponse({'error': 'غير مصرّح'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'POST مطلوب'}, status=405)

    try:
        body = _json.loads(request.body)
    except _json.JSONDecodeError:
        return JsonResponse({'error': 'بيانات غير صالحة'}, status=400)

    exam_id = body.get('exam_id')
    student_key = body.get('student_key', '')

    if not exam_id:
        return JsonResponse({'error': 'معرّف الاختبار مطلوب'}, status=400)

    teacher = get_teacher(request)
    exam = get_object_or_404(TeacherExam, id=exam_id, skill__created_by=teacher)

    from students.models import Student, RemedialExamAssignment

    student = None
    if student_key.startswith('sr_'):
        try:
            student = Student.objects.get(id=int(student_key[3:]))
        except (Student.DoesNotExist, ValueError):
            pass
    elif student_key.startswith('u_'):
        try:
            user_id = int(student_key[2:])
            target_user = User.objects.get(id=user_id)
            full_name = target_user.get_full_name() or target_user.username
            student = Student.objects.filter(full_name=full_name).first()
        except (User.DoesNotExist, ValueError):
            pass

    if not student:
        return JsonResponse({'error': 'الطالبة غير موجودة في سجل الطالبات — تأكدي من إضافتها عبر صفحة الاستيراد'}, status=404)

    obj, created = RemedialExamAssignment.objects.get_or_create(
        exam_id=exam_id, student=student
    )

    return JsonResponse({
        'ok': True,
        'created': created,
        'student': student.full_name,
        'exam': exam.skill.title if exam.skill else str(exam_id),
    })


@never_cache
@login_required
def teacher_stats_json(request):
    """
    يُرجع إحصاءات الداشبورد كـ JSON — مع دعم فلاتر الفصل ونوع المهارة.
    يُستدعى من JS كل 30 ثانية لتحديث الأرقام بدون إعادة تحميل الصفحة.
    """
    if not check_teacher(request):
        return JsonResponse({'error': 'unauthorized'}, status=403)

    teacher = get_teacher(request)
    if not teacher:
        return JsonResponse({'error': 'no teacher'}, status=403)

    # ─── فلاتر ───────────────────────────────────────────────
    f_classroom  = request.GET.get('classroom', '')
    f_skill_type = request.GET.get('skill_type', '')

    # فصول المعلمة المفلترة
    _teacher_classrooms = teacher.classrooms.all()
    if f_classroom and f_classroom != 'all':
        _filtered_classrooms = _teacher_classrooms.filter(name=f_classroom)
    else:
        _filtered_classrooms = _teacher_classrooms

    # طالبات الفصل المفلتر
    _cls_qs = Student.objects.filter(classroom__in=_filtered_classrooms)

    my_skills  = TeacherSkill.objects.filter(created_by=teacher)
    my_exams   = TeacherExam.objects.filter(skill__created_by=teacher)
    my_results = ExamResult.objects.filter(
        exam__skill__created_by=teacher
    ).exclude(student=request.user)

    # تطبيق فلتر نوع المهارة
    if f_skill_type and f_skill_type != 'all':
        type_map = {'qodrat': 'skill', 'tahsili': 'lesson'}
        mapped = type_map.get(f_skill_type, f_skill_type)
        my_skills  = my_skills.filter(content_type=mapped)
        my_exams   = my_exams.filter(skill__content_type=mapped)
        my_results = my_results.filter(exam__skill__content_type=mapped)

    # تطبيق فلتر الفصل على النتائج
    if f_classroom and f_classroom != 'all':
        _user_ids   = _cls_qs.filter(user__isnull=False).values_list('user_id', flat=True)
        _record_ids = _cls_qs.values_list('id', flat=True)
        my_results = my_results.filter(
            Q(student_id__in=_user_ids) | Q(student_record_id__in=_record_ids)
        )

    # تصنيف الطالبات
    from collections import defaultdict as _dd
    perf_map = _dd(lambda: {'total': 0, 'count': 0})
    for r in my_results.select_related('student', 'student_record'):
        key = f'sr_{r.student_record_id}' if r.student_record_id else f'u_{r.student_id}'
        if not key.endswith('_None'):
            perf_map[key]['total'] += float(r.percentage or 0)
            perf_map[key]['count'] += 1

    excellent = mid = weak = 0
    for v in perf_map.values():
        if v['count']:
            avg = v['total'] / v['count']
            if avg >= 90:
                excellent += 1
            elif avg >= 50:
                mid += 1
            else:
                weak += 1

    avg_score = int(round(my_results.aggregate(avg=Avg('percentage'))['avg'] or 0))
    trained   = my_results.values('student_id').distinct().count()

    return JsonResponse({
        'total_students':           _cls_qs.count(),
        'total_skills':             my_skills.filter(content_type='skill').count(),
        'total_lessons':            my_skills.filter(content_type='lesson').count(),
        'total_banks':              my_skills.filter(content_type='bank').count(),
        'active_skills':            my_skills.filter(is_active=True).count(),
        'total_exams':              my_exams.count(),
        'active_exams':             my_exams.filter(is_active=True, results__in=my_results).distinct().count(),
        'total_results':            my_results.count(),
        'passed_results':           my_results.filter(passed=True).count(),
        'avg_score':                avg_score,
        'trained_students':         trained,
        'students_count_excellent': excellent,
        'students_count_mid':       mid,
        'students_count_weak':      weak,
    })
